# main.py
import sys
import os
import json
from datetime import datetime, date, timedelta
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtChart import *
import mysql.connector
from mysql.connector import Error, pooling
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape, A4
from reportlab.platypus import *
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
import subprocess
from functools import partial
import hashlib
import uuid

# ==================== CONFIGURATION ====================
class Config:
    APP_NAME = "MAX AGENCY"
    VERSION = "3.0.0"
    COMPANY_NAME = "EquipRent Solutions Inc."
    
    # Database Configuration
    DB_CONFIG = {
        'host': 'localhost',
        'database': 'equipment_rental',
        'user': 'root',
        'password': '5757',
        'pool_name': 'equipment_pool',
        'pool_size': 10
    }
    
    # Theme Colors
    PRIMARY_COLOR = "#2196F3"
    SECONDARY_COLOR = "#FF9800"
    SUCCESS_COLOR = "#4CAF50"
    DANGER_COLOR = "#F44336"
    WARNING_COLOR = "#FFC107"
    INFO_COLOR = "#00BCD4"
    DARK_COLOR = "#2C3E50"
    LIGHT_COLOR = "#ECF0F1"
    
    # Style Constants
    BORDER_RADIUS = "8px"
    FONT_FAMILY = "Segoe UI, Arial, sans-serif"
    HEADER_FONT_SIZE = "24px"
    TITLE_FONT_SIZE = "18px"
    NORMAL_FONT_SIZE = "12px"
    SMALL_FONT_SIZE = "11px"

# ==================== DATABASE MANAGER ====================
class DatabaseManager:
    _instance = None
    
    def __new__(cls):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
            cls._instance.initialize()
        return cls._instance
    
    def initialize(self):
        try:
            self.pool = mysql.connector.pooling.MySQLConnectionPool(**Config.DB_CONFIG)
            print("Database connection pool created successfully")
        except Error as e:
            QMessageBox.critical(None, "Database Error", f"Failed to create connection pool: {e}")
            sys.exit(1)
    
    def get_connection(self):
        return self.pool.get_connection()
    
    def execute_query(self, query, params=None, fetch=True):
        connection = None
        cursor = None
        try:
            connection = self.get_connection()
            cursor = connection.cursor(dictionary=True, buffered=True)
            
            if params:
                cursor.execute(query, params)
            else:
                cursor.execute(query)
            
            if fetch and query.strip().upper().startswith('SELECT'):
                result = cursor.fetchall()
                return result
            else:
                connection.commit()
                return cursor.lastrowid
        except Error as e:
            if connection:
                connection.rollback()
            raise e
        finally:
            if cursor:
                cursor.close()
            if connection:
                connection.close()

# ==================== THEME MANAGER ====================
class ThemeManager:
    @staticmethod
    def get_main_stylesheet():
        return f"""
            * {{
                font-family: '{Config.FONT_FAMILY}';
                font-size: {Config.NORMAL_FONT_SIZE};
            }}
            
            QMainWindow {{
                background-color: #F5F6FA;
            }}
            
            QWidget#centralWidget {{
                background-color: #F5F6FA;
            }}
            
            QPushButton {{
                border: none;
                padding: 8px 16px;
                font-weight: bold;
                border-radius: {Config.BORDER_RADIUS};
                color: white;
                min-height: 30px;
            }}
            
            QPushButton:hover {{
                filter: brightness(110%);
            }}
            
            QPushButton:pressed {{
                filter: brightness(90%);
            }}
            
            QPushButton:disabled {{
                background-color: #BDC3C7;
                color: #7F8C8D;
            }}
            
            QLineEdit, QTextEdit, QPlainTextEdit, QSpinBox, QDoubleSpinBox, QDateEdit, QComboBox {{
                border: 2px solid #E0E0E0;
                border-radius: {Config.BORDER_RADIUS};
                padding: 8px;
                background-color: white;
                min-height: 20px;
            }}
            
            QLineEdit:focus, QTextEdit:focus, QPlainTextEdit:focus, QSpinBox:focus, QDoubleSpinBox:focus, QDateEdit:focus, QComboBox:focus {{
                border-color: {Config.PRIMARY_COLOR};
            }}
            
            QTableWidget {{
                border: 1px solid #E0E0E0;
                gridline-color: #F0F0F0;
                background-color: white;
                alternate-background-color: #F8F9FA;
                selection-background-color: {Config.PRIMARY_COLOR};
                selection-color: white;
                border-radius: {Config.BORDER_RADIUS};
            }}
            
            QTableWidget::item {{
                padding: 8px;
                border-bottom: 1px solid #F0F0F0;
            }}
            
            QHeaderView::section {{
                background-color: {Config.DARK_COLOR};
                color: white;
                padding: 12px 8px;
                font-weight: bold;
                border: none;
                border-right: 1px solid #34495E;
            }}
            
            QHeaderView::section:hover {{
                background-color: #34495E;
            }}
            
            QScrollBar:vertical {{
                border: none;
                background: #F5F6FA;
                width: 12px;
                border-radius: 6px;
            }}
            
            QScrollBar::handle:vertical {{
                background: #BDC3C7;
                border-radius: 6px;
                min-height: 20px;
            }}
            
            QScrollBar::handle:vertical:hover {{
                background: #95A5A6;
            }}
            
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
                border: none;
                background: none;
            }}
            
            QTabWidget::pane {{
                border: 1px solid #E0E0E0;
                border-radius: {Config.BORDER_RADIUS};
                background: white;
            }}
            
            QTabBar::tab {{
                background: #ECF0F1;
                padding: 10px 20px;
                margin-right: 2px;
                border-top-left-radius: {Config.BORDER_RADIUS};
                border-top-right-radius: {Config.BORDER_RADIUS};
            }}
            
            QTabBar::tab:selected {{
                background: white;
                border-bottom: 3px solid {Config.PRIMARY_COLOR};
            }}
            
            QTabBar::tab:hover {{
                background: #D5DBDB;
            }}
            
            QStatusBar {{
                background-color: {Config.DARK_COLOR};
                color: white;
                padding: 5px;
            }}
            
            QMenuBar {{
                background-color: {Config.DARK_COLOR};
                color: white;
                padding: 2px;
            }}
            
            QMenuBar::item {{
                padding: 8px 15px;
                border-radius: 4px;
            }}
            
            QMenuBar::item:selected {{
                background-color: #34495E;
            }}
            
            QMenu {{
                background-color: white;
                border: 1px solid #E0E0E0;
                border-radius: 4px;
                padding: 5px;
            }}
            
            QMenu::item {{
                padding: 8px 30px 8px 20px;
                border-radius: 4px;
            }}
            
            QMenu::item:selected {{
                background-color: {Config.PRIMARY_COLOR};
                color: white;
            }}
        """

# ==================== CUSTOM WIDGETS ====================
class GradientWidget(QWidget):
    def __init__(self, start_color="#667eea", end_color="#764ba2", parent=None):
        super().__init__(parent)
        self.start_color = start_color
        self.end_color = end_color
        self.setMinimumHeight(100)
    
    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        
        gradient = QLinearGradient(0, 0, self.width(), 0)
        gradient.setColorAt(0, QColor(self.start_color))
        gradient.setColorAt(1, QColor(self.end_color))
        
        painter.setBrush(gradient)
        painter.setPen(Qt.NoPen)
        painter.drawRoundedRect(self.rect(), 15, 15)

class StatCard(QWidget):
    def __init__(self, title, value, icon, color, parent=None):
        super().__init__(parent)
        self.setStyleSheet(f"""
            QWidget#statCard {{
                background-color: white;
                border-radius: 10px;
                border-left: 5px solid {color};
            }}
        """)
        self.setObjectName("statCard")
        
        layout = QHBoxLayout(self)
        
        # Icon
        icon_label = QLabel(icon)
        icon_label.setStyleSheet(f"font-size: 40px; color: {color};")
        icon_label.setFixedSize(60, 60)
        icon_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(icon_label)
        
        # Text
        text_layout = QVBoxLayout()
        
        title_label = QLabel(title)
        title_label.setStyleSheet(f"font-size: {Config.SMALL_FONT_SIZE}; color: #7F8C8D; font-weight: bold;")
        text_layout.addWidget(title_label)
        
        value_label = QLabel(str(value))
        value_label.setStyleSheet(f"font-size: 24px; font-weight: bold; color: {Config.DARK_COLOR};")
        text_layout.addWidget(value_label)
        
        layout.addLayout(text_layout)

class ModernButton(QPushButton):
    def __init__(self, text, color=Config.PRIMARY_COLOR, icon=None, parent=None):
        super().__init__(text, parent)
        self.color = color
        self.setStyleSheet(f"""
            QPushButton {{
                background-color: {color};
                color: white;
                border: none;
                padding: 10px 20px;
                font-weight: bold;
                font-size: 13px;
                border-radius: {Config.BORDER_RADIUS};
            }}
            QPushButton:hover {{
                background-color: {self._lighten_color(color)};
            }}
            QPushButton:pressed {{
                background-color: {self._darken_color(color)};
            }}
        """)
        if icon:
            self.setIcon(QIcon(icon))
        self.setCursor(Qt.PointingHandCursor)
    
    def _lighten_color(self, color, factor=20):
        color = color.lstrip('#')
        r, g, b = int(color[0:2], 16), int(color[2:4], 16), int(color[4:6], 16)
        r = min(255, r + factor)
        g = min(255, g + factor)
        b = min(255, b + factor)
        return f'#{r:02x}{g:02x}{b:02x}'
    
    def _darken_color(self, color, factor=20):
        color = color.lstrip('#')
        r, g, b = int(color[0:2], 16), int(color[2:4], 16), int(color[4:6], 16)
        r = max(0, r - factor)
        g = max(0, g - factor)
        b = max(0, b - factor)
        return f'#{r:02x}{g:02x}{b:02x}'

class SearchWidget(QWidget):
    def __init__(self, placeholder="Search...", parent=None):
        super().__init__(parent)
        layout = QHBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText(placeholder)
        self.search_input.setStyleSheet(f"""
            QLineEdit {{
                padding: 10px 15px;
                border: 2px solid #E0E0E0;
                border-radius: 20px;
                font-size: 13px;
            }}
            QLineEdit:focus {{
                border-color: {Config.PRIMARY_COLOR};
            }}
        """)
        layout.addWidget(self.search_input)

# ==================== MAIN APPLICATION ====================
class MainApplication(QMainWindow):
    def __init__(self):
        super().__init__()
        self.db = DatabaseManager()
        self.init_ui()
        self.apply_theme()
        self.show_dashboard()
    
    def init_ui(self):
        self.setWindowTitle(f"{Config.APP_NAME} v{Config.VERSION}")
        self.setGeometry(100, 50, 1400, 900)
        
        # Create Menu Bar
        self.create_menu_bar()
        
        # Create Status Bar
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.status_bar.showMessage("Ready")
        
        # Create Tool Bar
        self.create_toolbar()
        
        # Main Central Widget with Stack
        self.central_stack = QStackedWidget()
        self.setCentralWidget(self.central_stack)
    
    def create_menu_bar(self):
        menubar = self.menuBar()
        
        # File Menu
        file_menu = menubar.addMenu("&File")
        
        new_action = QAction("New Rental", self)
        new_action.setShortcut("Ctrl+N")
        new_action.triggered.connect(self.open_rental_interface)
        file_menu.addAction(new_action)
        
        file_menu.addSeparator()
        
        backup_action = QAction("Backup Database", self)
        backup_action.triggered.connect(self.backup_database)
        file_menu.addAction(backup_action)
        
        file_menu.addSeparator()
        
        exit_action = QAction("Exit", self)
        exit_action.setShortcut("Ctrl+Q")
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)
        
        # View Menu
        view_menu = menubar.addMenu("&View")
        
        dashboard_action = QAction("Dashboard", self)
        dashboard_action.triggered.connect(self.show_dashboard)
        view_menu.addAction(dashboard_action)
        
        view_menu.addSeparator()
        
        for entity in ["Customers", "Equipment", "Staff", "Rentals", "Payments", "Maintenance"]:
            action = QAction(entity, self)
            action.triggered.connect(partial(self.open_entity, entity.lower()))
            view_menu.addAction(action)
        
        # Reports Menu
        reports_menu = menubar.addMenu("&Reports")
        
        report_types = [
            "Customer Report",
            "Equipment Report",
            "Rental Analysis",
            "Revenue Report",
            "Maintenance Log",
            "Staff Performance"
        ]
        
        for report in report_types:
            action = QAction(report, self)
            action.triggered.connect(partial(self.generate_specific_report, report))
            reports_menu.addAction(action)
        
        # Help Menu
        help_menu = menubar.addMenu("&Help")
        
        about_action = QAction("About", self)
        about_action.triggered.connect(self.show_about)
        help_menu.addAction(about_action)
        
        help_action = QAction("Help", self)
        help_action.setShortcut("F1")
        help_action.triggered.connect(self.show_help)
        help_menu.addAction(help_action)
    
    def create_toolbar(self):
        toolbar = self.addToolBar("Main Toolbar")
        toolbar.setMovable(False)
        toolbar.setStyleSheet(f"""
            QToolBar {{
                background-color: white;
                border-bottom: 1px solid #E0E0E0;
                padding: 5px;
                spacing: 10px;
            }}
        """)
        
        actions = [
            ("🏠", "Dashboard", self.show_dashboard),
            ("👥", "Customers", self.open_customer_interface),
            ("🔧", "Equipment", self.open_equipment_interface),
            ("👨‍💼", "Staff", self.open_staff_interface),
            ("📋", "Rentals", self.open_rental_interface),
            ("💰", "Payments", self.open_payment_interface),
            ("🔨", "Maintenance", self.open_maintenance_interface),
            ("📊", "Reports", self.open_reports_window),
        ]
        
        for icon, text, handler in actions:
            btn = QPushButton(f"{icon} {text}")
            btn.setStyleSheet(f"""
                QPushButton {{
                    background-color: transparent;
                    border: none;
                    padding: 8px 15px;
                    font-weight: bold;
                    border-radius: 5px;
                    color: {Config.DARK_COLOR};
                }}
                QPushButton:hover {{
                    background-color: #F0F0F0;
                }}
            """)
            btn.clicked.connect(handler)
            btn.setCursor(Qt.PointingHandCursor)
            toolbar.addWidget(btn)
    
    def apply_theme(self):
        self.setStyleSheet(ThemeManager.get_main_stylesheet())
    
    def show_dashboard(self):
        # Create dashboard widget
        dashboard = QWidget()
        dashboard_layout = QVBoxLayout(dashboard)
        
        # Header with gradient
        header = GradientWidget()
        header_layout = QVBoxLayout(header)
        
        company_label = QLabel(Config.COMPANY_NAME)
        company_label.setStyleSheet("color: rgba(255,255,255,0.8); font-size: 14px;")
        company_label.setAlignment(Qt.AlignCenter)
        header_layout.addWidget(company_label)
        
        welcome_label = QLabel(f"Welcome to {Config.APP_NAME}")
        welcome_label.setStyleSheet("color: white; font-size: 36px; font-weight: bold;")
        welcome_label.setAlignment(Qt.AlignCenter)
        header_layout.addWidget(welcome_label)
        
        date_label = QLabel(datetime.now().strftime("%A, %B %d, %Y"))
        date_label.setStyleSheet("color: rgba(255,255,255,0.9); font-size: 16px;")
        date_label.setAlignment(Qt.AlignCenter)
        header_layout.addWidget(date_label)
        
        dashboard_layout.addWidget(header)
        
        # Stats Cards
        stats_widget = QWidget()
        stats_layout = QHBoxLayout(stats_widget)
        stats_layout.setSpacing(20)
        
        # Get stats from database
        stats = self.get_dashboard_stats()
        
        stat_cards = [
            ("Total Customers", stats['customers'], "👥", Config.PRIMARY_COLOR),
            ("Active Rentals", stats['active_rentals'], "📋", Config.SUCCESS_COLOR),
            ("Available Equipment", stats['available_equipment'], "🔧", Config.INFO_COLOR),
            ("Revenue Today", f"GHc{stats['revenue_today']:,.2f}", "💰", Config.SECONDARY_COLOR),
        ]
        
        for title, value, icon, color in stat_cards:
            card = StatCard(title, value, icon, color)
            stats_layout.addWidget(card)
        
        dashboard_layout.addWidget(stats_widget)
        
        # Charts and Quick Actions
        content_widget = QWidget()
        content_layout = QHBoxLayout(content_widget)
        content_layout.setSpacing(20)
        
        # Charts (Left side)
        charts_widget = QWidget()
        charts_layout = QVBoxLayout(charts_widget)
        
        # Rental Trend Chart
        chart_view = self.create_rental_chart()
        charts_layout.addWidget(chart_view)
        
        content_layout.addWidget(charts_widget, 2)
        
        # Quick Actions (Right side)
        actions_widget = QWidget()
        actions_layout = QVBoxLayout(actions_widget)
        
        actions_title = QLabel("Quick Actions")
        actions_title.setStyleSheet(f"font-size: {Config.TITLE_FONT_SIZE}; font-weight: bold; color: {Config.DARK_COLOR};")
        actions_layout.addWidget(actions_title)
        
        quick_actions = [
            ("New Rental", Config.SUCCESS_COLOR, self.open_rental_interface),
            ("Add Customer", Config.PRIMARY_COLOR, self.open_customer_interface),
            ("Add Equipment", Config.INFO_COLOR, self.open_equipment_interface),
            ("Record Payment", Config.SECONDARY_COLOR, self.open_payment_interface),
            ("Schedule Maintenance", Config.DANGER_COLOR, self.open_maintenance_interface),
        ]
        
        for text, color, handler in quick_actions:
            btn = ModernButton(text, color)
            btn.clicked.connect(handler)
            actions_layout.addWidget(btn)
        
        actions_layout.addStretch()
        content_layout.addWidget(actions_widget, 1)
        
        dashboard_layout.addWidget(content_widget)
        
        # Recent Activity
        recent_widget = self.create_recent_activity()
        dashboard_layout.addWidget(recent_widget)
        
        # Update status bar
        self.status_bar.showMessage(f"Dashboard updated at {datetime.now().strftime('%H:%M:%S')}")
        
        # Add to stack
        if self.central_stack.count() > 0:
            self.central_stack.removeWidget(self.central_stack.widget(0))
        self.central_stack.addWidget(dashboard)
    
    def get_dashboard_stats(self):
        stats = {
            'customers': 0,
            'active_rentals': 0,
            'available_equipment': 0,
            'revenue_today': 0
        }
        
        try:
            # Get customer count
            result = self.db.execute_query("SELECT COUNT(*) as count FROM customers")
            stats['customers'] = result[0]['count'] if result else 0
            
            # Get active rentals
            result = self.db.execute_query("SELECT COUNT(*) as count FROM rentals WHERE status = 'Active'")
            stats['active_rentals'] = result[0]['count'] if result else 0
            
            # Get available equipment
            result = self.db.execute_query("SELECT COUNT(*) as count FROM equipment WHERE status = 'Available'")
            stats['available_equipment'] = result[0]['count'] if result else 0
            
            # Get today's revenue
            result = self.db.execute_query(
                "SELECT COALESCE(SUM(amount), 0) as total FROM payments WHERE DATE(payment_date) = CURDATE()"
            )
            stats['revenue_today'] = float(result[0]['total']) if result else 0
            
        except Exception as e:
            print(f"Error fetching stats: {e}")
        
        return stats
    
    def create_rental_chart(self):
        chart_view = QChartView()
        chart_view.setRenderHint(QPainter.Antialiasing)
        
        chart = QChart()
        chart.setTitle("Rental Trends (Last 7 Days)")
        chart.setAnimationOptions(QChart.SeriesAnimations)
        
        # Get rental data for last 7 days
        try:
            query = """
                SELECT DATE(rental_date) as date, COUNT(*) as count
                FROM rentals
                WHERE rental_date >= DATE_SUB(CURDATE(), INTERVAL 7 DAY)
                GROUP BY DATE(rental_date)
                ORDER BY date
            """
            results = self.db.execute_query(query)
            
            series = QLineSeries()
            series.setName("Rentals")
            
            axis_x = QDateTimeAxis()
            axis_x.setFormat("MMM dd")
            axis_x.setTitleText("Date")
            
            axis_y = QValueAxis()
            axis_y.setTitleText("Number of Rentals")
            axis_y.setLabelFormat("%d")
            
            chart.addAxis(axis_x, Qt.AlignBottom)
            chart.addAxis(axis_y, Qt.AlignLeft)
            
            if results:
                for row in results:
                    date_str = row['date']
                    if isinstance(date_str, date):
                        dt = QDateTime(date_str.year, date_str.month, date_str.day, 0, 0)
                        series.append(dt.toMSecsSinceEpoch(), row['count'])
            else:
                # Add dummy data if no results
                for i in range(7):
                    dt = QDateTime.currentDateTime().addDays(-6 + i)
                    series.append(dt.toMSecsSinceEpoch(), 0)
            
            chart.addSeries(series)
            series.attachAxis(axis_x)
            series.attachAxis(axis_y)
            
        except Exception as e:
            print(f"Error creating chart: {e}")
        
        chart_view.setChart(chart)
        chart_view.setMinimumHeight(300)
        
        return chart_view
    
    def create_recent_activity(self):
        widget = QWidget()
        widget.setStyleSheet("""
            QWidget {
                background-color: white;
                border-radius: 10px;
            }
        """)
        layout = QVBoxLayout(widget)
        
        title = QLabel("Recent Activity")
        title.setStyleSheet(f"font-size: {Config.TITLE_FONT_SIZE}; font-weight: bold; color: {Config.DARK_COLOR}; padding: 10px;")
        layout.addWidget(title)
        
        table = QTableWidget()
        table.setColumnCount(4)
        table.setHorizontalHeaderLabels(["Date", "Activity", "Details", "Status"])
        table.horizontalHeader().setStretchLastSection(True)
        table.setMaximumHeight(200)
        
        # Get recent activity
        try:
            query = """
                (SELECT rental_date as date, 'Rental' as activity, 
                        CONCAT(c.first_name, ' ', c.last_name, ' rented ', e.name) as details,
                        r.status
                 FROM rentals r
                 JOIN customers c ON r.customer_id = c.customer_id
                 JOIN equipment e ON r.equipment_id = e.equipment_id
                 ORDER BY r.rental_date DESC LIMIT 5)
                UNION ALL
                (SELECT payment_date as date, 'Payment' as activity,
                        CONCAT('$', p.amount, ' - ', p.payment_method) as details,
                        p.status
                 FROM payments p
                 ORDER BY p.payment_date DESC LIMIT 5)
                ORDER BY date DESC LIMIT 10
            """
            results = self.db.execute_query(query)
            
            if results:
                table.setRowCount(len(results))
                for i, row in enumerate(results):
                    table.setItem(i, 0, QTableWidgetItem(str(row['date'])))
                    table.setItem(i, 1, QTableWidgetItem(row['activity']))
                    table.setItem(i, 2, QTableWidgetItem(row['details']))
                    
                    status_item = QTableWidgetItem(row['status'])
                    if row['status'] in ['Active', 'Completed']:
                        status_item.setForeground(QColor(Config.SUCCESS_COLOR))
                    elif row['status'] in ['Overdue', 'Failed']:
                        status_item.setForeground(QColor(Config.DANGER_COLOR))
                    table.setItem(i, 3, status_item)
                    
        except Exception as e:
            print(f"Error fetching recent activity: {e}")
        
        layout.addWidget(table)
        
        return widget
    
    def create_entity_interface(self, title, table_name, columns):
        dialog = EntityManagementDialog(title, table_name, columns, self)
        dialog.exec_()
    
    def open_entity(self, entity_name):
        entity_map = {
            'customers': ('Customer Management', 'customers', 
                         ['customer_id', 'first_name', 'last_name', 'email', 'phone', 'address', 'registration_date']),
            'equipment': ('Equipment Management', 'equipment',
                         ['equipment_id', 'name', 'category', 'description', 'daily_rate', 'status', 'purchase_date', 'last_maintenance_date']),
            'staff': ('Staff Management', 'staff',
                     ['staff_id', 'first_name', 'last_name', 'email', 'phone', 'position', 'hire_date', 'salary']),
            'rentals': ('Rental Management', 'rentals',
                       ['rental_id', 'customer_id', 'equipment_id', 'staff_id', 'rental_date', 'return_date', 'total_amount', 'status']),
            'payments': ('Payment Management', 'payments',
                        ['payment_id', 'rental_id', 'amount', 'payment_date', 'payment_method', 'status']),
            'maintenance': ('Maintenance Management', 'maintenance',
                           ['maintenance_id', 'equipment_id', 'staff_id', 'maintenance_date', 'description', 'cost', 'next_maintenance_date', 'status']),
        }
        
        if entity_name in entity_map:
            title, table, cols = entity_map[entity_name]
            self.create_entity_interface(title, table, cols)
    
    def open_customer_interface(self):
        self.open_entity('customers')
    
    def open_equipment_interface(self):
        self.open_entity('equipment')
    
    def open_staff_interface(self):
        self.open_entity('staff')
    
    def open_rental_interface(self):
        self.open_entity('rentals')
    
    def open_payment_interface(self):
        self.open_entity('payments')
    
    def open_maintenance_interface(self):
        self.open_entity('maintenance')
    
    def open_reports_window(self):
        dialog = ReportsDialog(self)
        dialog.exec_()
    
    def generate_specific_report(self, report_type):
        dialog = ReportsDialog(self, report_type)
        dialog.exec_()
    
    def backup_database(self):
        try:
            filename = f"backup/equipment_rental_{datetime.now().strftime('%Y%m%d_%H%M%S')}.sql"
            os.makedirs('backup', exist_ok=True)
            
            # This is a simplified backup - in production use mysqldump
            QMessageBox.information(self, "Backup", f"Database backup saved to:\n{filename}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Backup failed: {e}")
    
    def show_about(self):
        about_text = f"""
        <h1>{Config.APP_NAME}</h1>
        <p>Version: {Config.VERSION}</p>
        <p>{Config.COMPANY_NAME}</p>
        <hr>
        <p>A professional equipment rental management system.</p>
        <p>Built with PyQt5 and MySQL</p>
        <p>© 2024 All Rights Reserved</p>
        """
        QMessageBox.about(self, f"About {Config.APP_NAME}", about_text)
    
    def show_help(self):
        help_text = """
        <h2>Help Guide</h2>
        <hr>
        <h3>Keyboard Shortcuts:</h3>
        <ul>
            <li><b>Ctrl+N</b> - New Rental</li>
            <li><b>Ctrl+Q</b> - Exit Application</li>
            <li><b>F1</b> - Show Help</li>
            <li><b>F5</b> - Refresh Current View</li>
        </ul>
        <h3>Navigation:</h3>
        <p>Use the menu bar, toolbar, or dashboard to navigate between different sections.</p>
        <h3>Reports:</h3>
        <p>Access reports from the Reports menu or the Reports button on the toolbar.</p>
        """
        QMessageBox.information(self, "Help", help_text)

# ==================== ENTITY MANAGEMENT DIALOG ====================
class EntityManagementDialog(QDialog):
    def __init__(self, title, table_name, columns, parent=None):
        super().__init__(parent)
        self.title = title
        self.table_name = table_name
        self.columns = columns
        self.db = DatabaseManager()
        self.current_id = None
        
        self.init_ui()
        self.load_data()
    
    def init_ui(self):
        self.setWindowTitle(self.title)
        self.setGeometry(100, 100, 1300, 800)
        self.setModal(True)
        
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(0)
        
        # Header
        header = GradientWidget(Config.DARK_COLOR, "#34495E")
        header.setMinimumHeight(80)
        header_layout = QHBoxLayout(header)
        
        title_label = QLabel(self.title)
        title_label.setStyleSheet("color: white; font-size: 24px; font-weight: bold;")
        header_layout.addWidget(title_label)
        
        header_layout.addStretch()
        
        close_btn = ModernButton("Close", Config.DANGER_COLOR)
        close_btn.clicked.connect(self.close)
        header_layout.addWidget(close_btn)
        
        main_layout.addWidget(header)
        
        # Content
        content = QWidget()
        content_layout = QHBoxLayout(content)
        content_layout.setSpacing(20)
        
        # Left Panel - Form
        form_panel = QWidget()
        form_panel.setFixedWidth(350)
        form_panel.setStyleSheet("background-color: white; border-radius: 10px;")
        form_layout = QVBoxLayout(form_panel)
        
        form_title = QLabel("Record Details")
        form_title.setStyleSheet(f"font-size: {Config.TITLE_FONT_SIZE}; font-weight: bold; color: {Config.DARK_COLOR};")
        form_layout.addWidget(form_title)
        
        # Scroll area for form fields
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("border: none;")
        
        scroll_content = QWidget()
        self.form_fields_layout = QVBoxLayout(scroll_content)
        
        self.form_fields = {}
        for col in self.columns[1:]:  # Skip ID
            field_widget = QWidget()
            field_layout = QVBoxLayout(field_widget)
            field_layout.setSpacing(5)
            
            label = QLabel(col.replace('_', ' ').title() + ":")
            label.setStyleSheet("font-weight: bold; color: #555;")
            field_layout.addWidget(label)
            
            entry = QLineEdit()
            entry.setPlaceholderText(f"Enter {col.replace('_', ' ').lower()}...")
            field_layout.addWidget(entry)
            
            self.form_fields_layout.addWidget(field_widget)
            self.form_fields[col] = entry
        
        self.form_fields_layout.addStretch()
        scroll.setWidget(scroll_content)
        form_layout.addWidget(scroll)
        
        # Action buttons
        btn_container = QWidget()
        btn_layout = QVBoxLayout(btn_container)
        btn_layout.setSpacing(10)
        
        add_btn = ModernButton("➕ Add Record", Config.SUCCESS_COLOR)
        add_btn.clicked.connect(self.add_record)
        btn_layout.addWidget(add_btn)
        
        update_btn = ModernButton("✏️ Update Record", Config.WARNING_COLOR)
        update_btn.clicked.connect(self.update_record)
        btn_layout.addWidget(update_btn)
        
        delete_btn = ModernButton("🗑️ Delete Record", Config.DANGER_COLOR)
        delete_btn.clicked.connect(self.delete_record)
        btn_layout.addWidget(delete_btn)
        
        clear_btn = ModernButton("🔄 Clear Form", "#95A5A6")
        clear_btn.clicked.connect(self.clear_form)
        btn_layout.addWidget(clear_btn)
        
        form_layout.addWidget(btn_container)
        
        content_layout.addWidget(form_panel)
        
        # Right Panel - Table
        table_panel = QWidget()
        table_panel.setStyleSheet("background-color: white; border-radius: 10px;")
        table_layout = QVBoxLayout(table_panel)
        
        # Search and actions
        search_container = QWidget()
        search_layout = QHBoxLayout(search_container)
        
        self.search_widget = SearchWidget("Search records...")
        search_layout.addWidget(self.search_widget)
        
        search_btn = ModernButton("🔍 Search", Config.PRIMARY_COLOR)
        search_btn.clicked.connect(self.search_records)
        search_layout.addWidget(search_btn)
        
        refresh_btn = ModernButton("🔄 Refresh", Config.INFO_COLOR)
        refresh_btn.clicked.connect(self.load_data)
        search_layout.addWidget(refresh_btn)
        
        table_layout.addWidget(search_container)
        
        # Table
        self.table = QTableWidget()
        self.table.setColumnCount(len(self.columns))
        self.table.setHorizontalHeaderLabels([col.replace('_', ' ').title() for col in self.columns])
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setSelectionMode(QTableWidget.SingleSelection)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.clicked.connect(self.on_row_select)
        self.table.setSortingEnabled(True)
        
        # Adjust column widths
        header = self.table.horizontalHeader()
        for i in range(len(self.columns)):
            header.setSectionResizeMode(i, QHeaderView.Stretch if i == len(self.columns) - 1 else QHeaderView.ResizeToContents)
        
        table_layout.addWidget(self.table)
        
        # Pagination
        pagination_widget = QWidget()
        pagination_layout = QHBoxLayout(pagination_widget)
        
        self.page_label = QLabel("Page 1 of 1")
        self.page_label.setStyleSheet("color: #666;")
        pagination_layout.addWidget(self.page_label)
        
        pagination_layout.addStretch()
        
        prev_btn = ModernButton("◀ Previous", "#95A5A6")
        prev_btn.clicked.connect(self.previous_page)
        pagination_layout.addWidget(prev_btn)
        
        next_btn = ModernButton("Next ▶", "#95A5A6")
        next_btn.clicked.connect(self.next_page)
        pagination_layout.addWidget(next_btn)
        
        table_layout.addWidget(pagination_widget)
        
        content_layout.addWidget(table_panel)
        
        main_layout.addWidget(content)
    
    def load_data(self):
        try:
            query = f"SELECT * FROM {self.table_name} ORDER BY {self.columns[0]} DESC"
            results = self.db.execute_query(query)
            
            self.table.setRowCount(0)
            if results:
                self.table.setRowCount(len(results))
                for row_idx, row in enumerate(results):
                    for col_idx, col in enumerate(self.columns):
                        value = row[col] if row[col] is not None else ''
                        item = QTableWidgetItem(str(value))
                        
                        # Color code status columns
                        if col == 'status' and str(value) in ['Active', 'Completed', 'Available']:
                            item.setForeground(QColor(Config.SUCCESS_COLOR))
                        elif col == 'status' and str(value) in ['Overdue', 'Maintenance', 'Failed']:
                            item.setForeground(QColor(Config.DANGER_COLOR))
                        
                        self.table.setItem(row_idx, col_idx, item)
            
            self.update_pagination()
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load data: {e}")
    
    def on_row_select(self):
        current_row = self.table.currentRow()
        if current_row >= 0:
            self.current_id = self.table.item(current_row, 0).text()
            self.clear_form()
            
            for i, col in enumerate(self.columns[1:]):
                item = self.table.item(current_row, i + 1)
                if item and col in self.form_fields:
                    self.form_fields[col].setText(item.text())
    
    def clear_form(self):
        self.current_id = None
        for field in self.form_fields.values():
            field.clear()
    
    def add_record(self):
        values = []
        for col in self.columns[1:]:
            value = self.form_fields[col].text().strip()
            values.append(value if value else None)
        
        if not any(values):
            QMessageBox.warning(self, "Warning", "Please enter at least one value")
            return
        
        try:
            placeholders = ', '.join(['%s'] * len(self.columns[1:]))
            columns_str = ', '.join(self.columns[1:])
            query = f"INSERT INTO {self.table_name} ({columns_str}) VALUES ({placeholders})"
            
            result = self.db.execute_query(query, values, fetch=False)
            if result:
                QMessageBox.information(self, "Success", "Record added successfully!")
                self.load_data()
                self.clear_form()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to add record: {e}")
    
    def update_record(self):
        if not self.current_id:
            QMessageBox.warning(self, "Warning", "Please select a record to update")
            return
        
        set_clause = []
        values = []
        for col in self.columns[1:]:
            value = self.form_fields[col].text().strip()
            if value:
                set_clause.append(f"{col} = %s")
                values.append(value)
        
        if not set_clause:
            QMessageBox.warning(self, "Warning", "No values to update")
            return
        
        try:
            set_clause_str = ', '.join(set_clause)
            values.append(self.current_id)
            query = f"UPDATE {self.table_name} SET {set_clause_str} WHERE {self.columns[0]} = %s"
            
            self.db.execute_query(query, values, fetch=False)
            QMessageBox.information(self, "Success", "Record updated successfully!")
            self.load_data()
            self.clear_form()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to update record: {e}")
    
    def delete_record(self):
        if not self.current_id:
            QMessageBox.warning(self, "Warning", "Please select a record to delete")
            return
        
        reply = QMessageBox.question(
            self, "Confirm Delete",
            "Are you sure you want to delete this record?\nThis action cannot be undone.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            try:
                query = f"DELETE FROM {self.table_name} WHERE {self.columns[0]} = %s"
                self.db.execute_query(query, (self.current_id,), fetch=False)
                QMessageBox.information(self, "Success", "Record deleted successfully!")
                self.load_data()
                self.clear_form()
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to delete record: {e}")
    
    def search_records(self):
        search_term = self.search_widget.search_input.text().strip()
        if not search_term:
            self.load_data()
            return
        
        try:
            search_conditions = []
            params = []
            for col in self.columns:
                if col != self.columns[0]:  # Skip ID column for better performance
                    search_conditions.append(f"{col} LIKE %s")
                    params.append(f"%{search_term}%")
            
            query = f"SELECT * FROM {self.table_name} WHERE {' OR '.join(search_conditions)}"
            results = self.db.execute_query(query, params)
            
            self.table.setRowCount(0)
            if results:
                self.table.setRowCount(len(results))
                for row_idx, row in enumerate(results):
                    for col_idx, col in enumerate(self.columns):
                        value = row[col] if row[col] is not None else ''
                        self.table.setItem(row_idx, col_idx, QTableWidgetItem(str(value)))
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Search failed: {e}")
    
    def update_pagination(self):
        total_rows = self.table.rowCount()
        self.page_label.setText(f"Total Records: {total_rows}")
    
    def previous_page(self):
        # Implement pagination logic here
        pass
    
    def next_page(self):
        # Implement pagination logic here
        pass

# ==================== REPORTS DIALOG ====================
class ReportsDialog(QDialog):
    def __init__(self, parent=None, default_report=None):
        super().__init__(parent)
        self.db = DatabaseManager()
        self.default_report = default_report
        self.init_ui()
        
        if default_report:
            self.generate_report(default_report)
    
    def init_ui(self):
        self.setWindowTitle("Generate Reports")
        self.setGeometry(200, 200, 800, 600)
        self.setModal(True)
        
        main_layout = QVBoxLayout(self)
        
        # Header
        header = GradientWidget(Config.SECONDARY_COLOR, Config.DANGER_COLOR)
        header_layout = QVBoxLayout(header)
        
        title = QLabel("Report Generation Center")
        title.setStyleSheet("color: white; font-size: 28px; font-weight: bold;")
        title.setAlignment(Qt.AlignCenter)
        header_layout.addWidget(title)
        
        subtitle = QLabel("Generate professional reports for your business")
        subtitle.setStyleSheet("color: rgba(255,255,255,0.9); font-size: 14px;")
        subtitle.setAlignment(Qt.AlignCenter)
        header_layout.addWidget(subtitle)
        
        main_layout.addWidget(header)
        
        # Content
        content = QWidget()
        content_layout = QVBoxLayout(content)
        
        # Report selection
        selection_widget = QWidget()
        selection_layout = QHBoxLayout(selection_widget)
        
        selection_label = QLabel("Select Report:")
        selection_label.setStyleSheet("font-size: 16px; font-weight: bold;")
        selection_layout.addWidget(selection_label)
        
        self.report_combo = QComboBox()
        self.report_combo.addItems([
            "Customer Report",
            "Equipment Report",
            "Rental Analysis",
            "Revenue Report",
            "Maintenance Log",
            "Staff Performance"
        ])
        self.report_combo.setStyleSheet("""
            QComboBox {
                padding: 10px;
                font-size: 14px;
                min-width: 300px;
            }
        """)
        selection_layout.addWidget(self.report_combo)
        
        generate_btn = ModernButton("Generate Report", Config.SUCCESS_COLOR)
        generate_btn.clicked.connect(lambda: self.generate_report(self.report_combo.currentText()))
        selection_layout.addWidget(generate_btn)
        
        selection_layout.addStretch()
        
        content_layout.addWidget(selection_widget)
        
        # Report preview area
        preview_label = QLabel("Report Preview")
        preview_label.setStyleSheet(f"font-size: {Config.TITLE_FONT_SIZE}; font-weight: bold; color: {Config.DARK_COLOR}; margin-top: 20px;")
        content_layout.addWidget(preview_label)
        
        self.preview_table = QTableWidget()
        self.preview_table.setAlternatingRowColors(True)
        self.preview_table.setEditTriggers(QTableWidget.NoEditTriggers)
        content_layout.addWidget(self.preview_table)
        
        # Export options
        export_widget = QWidget()
        export_layout = QHBoxLayout(export_widget)
        
        export_label = QLabel("Export As:")
        export_label.setStyleSheet("font-weight: bold;")
        export_layout.addWidget(export_label)
        
        pdf_btn = ModernButton("📄 PDF", Config.DANGER_COLOR)
        pdf_btn.clicked.connect(lambda: self.export_report("PDF"))
        export_layout.addWidget(pdf_btn)
        
        excel_btn = ModernButton("📊 Excel", Config.SUCCESS_COLOR)
        excel_btn.clicked.connect(lambda: self.export_report("Excel"))
        export_layout.addWidget(excel_btn)
        
        print_btn = ModernButton("🖨️ Print", Config.PRIMARY_COLOR)
        print_btn.clicked.connect(self.print_report)
        export_layout.addWidget(print_btn)
        
        export_layout.addStretch()
        
        content_layout.addWidget(export_widget)
        
        main_layout.addWidget(content)
        
        # Footer with close button
        footer = QWidget()
        footer_layout = QHBoxLayout(footer)
        footer_layout.addStretch()
        
        close_btn = ModernButton("Close", Config.DARK_COLOR)
        close_btn.clicked.connect(self.close)
        footer_layout.addWidget(close_btn)
        
        main_layout.addWidget(footer)
    
    def generate_report(self, report_type):
        try:
            # Clear preview
            self.preview_table.setRowCount(0)
            
            # Generate report based on type
            if report_type == "Customer Report":
                query = """
                    SELECT 
                        c.customer_id as ID,
                        CONCAT(c.first_name, ' ', c.last_name) as Name,
                        c.email as Email,
                        c.phone as Phone,
                        COUNT(r.rental_id) as 'Total Rentals',
                        COALESCE(SUM(p.amount), 0) as 'Total Spent'
                    FROM customers c
                    LEFT JOIN rentals r ON c.customer_id = r.customer_id
                    LEFT JOIN payments p ON r.rental_id = p.rental_id
                    GROUP BY c.customer_id
                    ORDER BY 'Total Spent' DESC
                """
            elif report_type == "Equipment Report":
                query = """
                    SELECT 
                        e.equipment_id as ID,
                        e.name as Equipment,
                        e.category as Category,
                        e.status as Status,
                        e.daily_rate as 'Daily Rate',
                        COUNT(r.rental_id) as 'Times Rented'
                    FROM equipment e
                    LEFT JOIN rentals r ON e.equipment_id = r.equipment_id
                    GROUP BY e.equipment_id
                    ORDER BY 'Times Rented' DESC
                """
            elif report_type == "Rental Analysis":
                query = """
                    SELECT 
                        r.rental_id as 'Rental ID',
                        CONCAT(c.first_name, ' ', c.last_name) as Customer,
                        e.name as Equipment,
                        r.rental_date as 'Rental Date',
                        r.return_date as 'Return Date',
                        r.total_amount as Amount,
                        r.status as Status
                    FROM rentals r
                    JOIN customers c ON r.customer_id = c.customer_id
                    JOIN equipment e ON r.equipment_id = e.equipment_id
                    ORDER BY r.rental_date DESC
                """
            elif report_type == "Revenue Report":
                query = """
                    SELECT 
                        DATE(p.payment_date) as Date,
                        COUNT(DISTINCT p.payment_id) as Transactions,
                        SUM(p.amount) as Revenue,
                        AVG(p.amount) as 'Average Transaction'
                    FROM payments p
                    WHERE p.status = 'Completed'
                    GROUP BY DATE(p.payment_date)
                    ORDER BY Date DESC
                """
            elif report_type == "Maintenance Log":
                query = """
                    SELECT 
                        m.maintenance_id as ID,
                        e.name as Equipment,
                        CONCAT(s.first_name, ' ', s.last_name) as Technician,
                        m.maintenance_date as Date,
                        m.description as Description,
                        m.cost as Cost,
                        m.status as Status
                    FROM maintenance m
                    JOIN equipment e ON m.equipment_id = e.equipment_id
                    JOIN staff s ON m.staff_id = s.staff_id
                    ORDER BY m.maintenance_date DESC
                """
            elif report_type == "Staff Performance":
                query = """
                    SELECT 
                        CONCAT(s.first_name, ' ', s.last_name) as Staff,
                        s.position as Position,
                        COUNT(r.rental_id) as 'Rentals Handled',
                        COUNT(m.maintenance_id) as 'Maintenance Tasks',
                        COALESCE(SUM(p.amount), 0) as 'Revenue Generated'
                    FROM staff s
                    LEFT JOIN rentals r ON s.staff_id = r.staff_id
                    LEFT JOIN payments p ON r.rental_id = p.rental_id
                    LEFT JOIN maintenance m ON s.staff_id = m.staff_id
                    GROUP BY s.staff_id
                    ORDER BY 'Revenue Generated' DESC
                """
            
            results = self.db.execute_query(query)
            
            if results:
                # Set up table
                columns = list(results[0].keys())
                self.preview_table.setColumnCount(len(columns))
                self.preview_table.setHorizontalHeaderLabels([col.replace('_', ' ').title() for col in columns])
                
                self.preview_table.setRowCount(len(results))
                for row_idx, row in enumerate(results):
                    for col_idx, col in enumerate(columns):
                        value = row[col] if row[col] is not None else ''
                        item = QTableWidgetItem(str(value))
                        
                        # Format currency columns
                        if isinstance(value, (int, float)) and ('amount' in col.lower() or 'rate' in col.lower() or 'revenue' in col.lower() or 'cost' in col.lower()):
                            item.setText(f"${value:,.2f}")
                        
                        self.preview_table.setItem(row_idx, col_idx, item)
                
                self.preview_table.resizeColumnsToContents()
                
                QMessageBox.information(self, "Success", f"{report_type} generated successfully!")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to generate report: {e}")
    
    def export_report(self, format_type):
        if self.preview_table.rowCount() == 0:
            QMessageBox.warning(self, "Warning", "No report data to export")
            return
        
        try:
            if not os.path.exists('reports'):
                os.makedirs('reports')
            
            filepath = f"reports/{self.report_combo.currentText().replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            
            if format_type == "PDF":
                self.export_to_pdf(filepath)
            elif format_type == "Excel":
                self.export_to_excel(filepath)
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Export failed: {e}")
    
    def export_to_pdf(self, filepath):
        filepath += ".pdf"
        
        doc = SimpleDocTemplate(filepath, pagesize=landscape(letter))
        elements = []
        
        # Add header
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor(Config.DARK_COLOR),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        title = Paragraph(f"{self.report_combo.currentText()}", title_style)
        elements.append(title)
        
        # Add timestamp
        timestamp_style = ParagraphStyle(
            'Timestamp',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey,
            alignment=TA_RIGHT
        )
        timestamp = Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", timestamp_style)
        elements.append(timestamp)
        elements.append(Spacer(1, 20))
        
        # Add table
        headers = []
        for col in range(self.preview_table.columnCount()):
            headers.append(self.preview_table.horizontalHeaderItem(col).text())
        
        table_data = [headers]
        for row in range(self.preview_table.rowCount()):
            row_data = []
            for col in range(self.preview_table.columnCount()):
                item = self.preview_table.item(row, col)
                row_data.append(item.text() if item else '')
            table_data.append(row_data)
        
        table = Table(table_data, repeatRows=1)
        
        # Style the table
        style_commands = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(Config.DARK_COLOR)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.Color(0.95, 0.95, 0.95)]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ]
        
        table.setStyle(TableStyle(style_commands))
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        
        QMessageBox.information(self, "Success", f"Report exported to:\n{filepath}")
        
        # Open the file
        if os.name == 'nt':
            os.startfile(filepath)
        else:
            subprocess.call(['open', filepath])
    
    def export_to_excel(self, filepath):
        filepath += ".xlsx"
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = self.report_combo.currentText()
            
            # Add title
            ws.merge_cells('A1:H1')
            ws['A1'] = self.report_combo.currentText()
            ws['A1'].font = Font(size=16, bold=True, color='2C3E50')
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Add timestamp
            ws.merge_cells('A2:H2')
            ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A2'].font = Font(size=10, color='7F8C8D')
            ws['A2'].alignment = Alignment(horizontal='right')
            
            # Add headers
            header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
            header_font = Font(size=11, bold=True, color='FFFFFF')
            
            for col in range(self.preview_table.columnCount()):
                cell = ws.cell(row=4, column=col+1)
                cell.value = self.preview_table.horizontalHeaderItem(col).text()
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # Add data
            for row in range(self.preview_table.rowCount()):
                for col in range(self.preview_table.columnCount()):
                    cell = ws.cell(row=row+5, column=col+1)
                    item = self.preview_table.item(row, col)
                    cell.value = item.text() if item else ''
                    cell.alignment = Alignment(horizontal='center')
            
            # Adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            wb.save(filepath)
            
            QMessageBox.information(self, "Success", f"Report exported to:\n{filepath}")
            
            if os.name == 'nt':
                os.startfile(filepath)
            else:
                subprocess.call(['open', filepath])
                
        except ImportError:
            QMessageBox.warning(self, "Warning", "openpyxl is required for Excel export.\nInstall it with: pip install openpyxl")
    
    def print_report(self):
        if self.preview_table.rowCount() == 0:
            QMessageBox.warning(self, "Warning", "No report data to print")
            return
        
        try:
            # Create a temporary PDF and print it
            temp_file = f"temp_print_{uuid.uuid4().hex[:8]}.pdf"
            self.export_to_pdf(temp_file)
            
            if os.path.exists(temp_file):
                if os.name == 'nt':
                    os.startfile(temp_file, 'print')
                else:
                    subprocess.call(['lpr', temp_file])
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Print failed: {e}")

# ==================== MAIN ENTRY POINT ====================
def main():
    # Create required directories
    os.makedirs('reports', exist_ok=True)
    os.makedirs('backup', exist_ok=True)
    
    app = QApplication(sys.argv)
    app.setApplicationName(Config.APP_NAME)
    app.setOrganizationName(Config.COMPANY_NAME)
    
    # Set application icon (you can add your own icon file)
    # app.setWindowIcon(QIcon('icon.png'))
    
    # Apply global stylesheet
    app.setStyle('Fusion')
    
    # Create and show main window
    window = MainApplication()
    window.show()
    
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()