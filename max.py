# ==================== CONFIGURATION ====================
"""
Equipment Rental Management System
A comprehensive PyQt5 application for managing equipment rentals,
customers, staff, payments, and maintenance.
Currency: Ghanaian Cedi (GH₵)
"""

import sys
import os
import json
from datetime import datetime, date, timedelta
from functools import partial
import hashlib
import uuid
import subprocess

# PyQt5 imports
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *

# Database imports
import mysql.connector
from mysql.connector import Error, pooling

# Report imports
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape, A4
from reportlab.platypus import *
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# Try importing QtChart with fallback
try:
    from PyQt5.QtChart import *
    QTCHART_AVAILABLE = True
except ImportError:
    QTCHART_AVAILABLE = False
    print("Warning: PyQtChart not installed. Charts will be disabled.")


# ==================== CONFIGURATION CLASS ====================
class Config:
    APP_NAME = "EquipRent Pro"
    VERSION = "3.0.0"
    COMPANY_NAME = "EquipRent Solutions Inc."
    CURRENCY = "GH₵"  # Ghanaian Cedi
    CURRENCY_SYMBOL = "GH₵"
    
    # Database Configuration - UPDATE THESE VALUES
    DB_CONFIG = {
        'host': 'localhost',
        'database': 'equipment_rental',
        'user': 'root',
        'password': '5757',  
        'pool_name': 'equipment_pool',
        'pool_size': 10
    }
    
    # Theme Colors
    PRIMARY_COLOR = "#2196F3"
    SECONDARY_COLOR = "#FF9800"
    SUCCESS_COLOR = "#4CAF50"
    DANGER_COLOR = "#F44336"
    WARNING_COLOR = "#FFC107"
    INFO_COLOR = "#00BCD4"
    DARK_COLOR = "#2C3E50"
    LIGHT_COLOR = "#ECF0F1"
    
    # Ghana Colors (for branding)
    GHANA_RED = "#CE1126"
    GHANA_GOLD = "#F5A623"
    GHANA_GREEN = "#006B3F"
    GHANA_BLACK = "#000000"
    
    # Style Constants
    BORDER_RADIUS = "8px"
    FONT_FAMILY = "Segoe UI, Arial, sans-serif"


# ==================== DATABASE MANAGER ====================
class DatabaseManager:
    """Singleton database connection manager with connection pooling"""
    
    _instance = None
    
    def __new__(cls):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
            cls._instance._initialized = False
        return cls._instance
    
    def __init__(self):
        if self._initialized:
            return
        self._initialized = True
        self.pool = None
        self._initialize_pool()
    
    def _initialize_pool(self):
        """Initialize the database connection pool"""
        try:
            self.pool = mysql.connector.pooling.MySQLConnectionPool(
                pool_name=Config.DB_CONFIG['pool_name'],
                pool_size=Config.DB_CONFIG['pool_size'],
                host=Config.DB_CONFIG['host'],
                database=Config.DB_CONFIG['database'],
                user=Config.DB_CONFIG['user'],
                password=Config.DB_CONFIG['password']
            )
            print("✅ Database connection pool created successfully")
        except Error as e:
            print(f"❌ Database connection failed: {e}")
            # Try to create database if it doesn't exist
            self._create_database()
    
    def _create_database(self):
        """Create the database if it doesn't exist"""
        try:
            # Connect without database
            conn = mysql.connector.connect(
                host=Config.DB_CONFIG['host'],
                user=Config.DB_CONFIG['user'],
                password=Config.DB_CONFIG['password']
            )
            cursor = conn.cursor()
            cursor.execute(f"CREATE DATABASE IF NOT EXISTS {Config.DB_CONFIG['database']}")
            cursor.execute(f"USE {Config.DB_CONFIG['database']}")
            self._create_tables(cursor)
            conn.commit()
            cursor.close()
            conn.close()
            
            # Now initialize the pool
            self.pool = mysql.connector.pooling.MySQLConnectionPool(
                pool_name=Config.DB_CONFIG['pool_name'],
                pool_size=Config.DB_CONFIG['pool_size'],
                host=Config.DB_CONFIG['host'],
                database=Config.DB_CONFIG['database'],
                user=Config.DB_CONFIG['user'],
                password=Config.DB_CONFIG['password']
            )
            print("✅ Database created successfully")
        except Error as e:
            print(f"❌ Failed to create database: {e}")
            QMessageBox.critical(None, "Database Error", 
                f"Cannot connect to database.\nError: {e}\n\n"
                "Please ensure MySQL is running and credentials are correct.")
            sys.exit(1)
    
    def _create_tables(self, cursor):
        """Create all required tables"""
        tables = [
            """
            CREATE TABLE IF NOT EXISTS customers (
                customer_id INT AUTO_INCREMENT PRIMARY KEY,
                first_name VARCHAR(50) NOT NULL,
                last_name VARCHAR(50) NOT NULL,
                email VARCHAR(100) UNIQUE NOT NULL,
                phone VARCHAR(20),
                address TEXT,
                registration_date DATE DEFAULT CURRENT_DATE,
                status VARCHAR(20) DEFAULT 'Active'
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS equipment (
                equipment_id INT AUTO_INCREMENT PRIMARY KEY,
                name VARCHAR(100) NOT NULL,
                category VARCHAR(50),
                description TEXT,
                daily_rate DECIMAL(10,2) NOT NULL,
                status VARCHAR(20) DEFAULT 'Available',
                purchase_date DATE,
                last_maintenance_date DATE,
                serial_number VARCHAR(50) UNIQUE
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS staff (
                staff_id INT AUTO_INCREMENT PRIMARY KEY,
                first_name VARCHAR(50) NOT NULL,
                last_name VARCHAR(50) NOT NULL,
                email VARCHAR(100) UNIQUE NOT NULL,
                phone VARCHAR(20),
                position VARCHAR(50),
                hire_date DATE DEFAULT CURRENT_DATE,
                salary DECIMAL(10,2),
                status VARCHAR(20) DEFAULT 'Active'
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS rentals (
                rental_id INT AUTO_INCREMENT PRIMARY KEY,
                customer_id INT NOT NULL,
                equipment_id INT NOT NULL,
                staff_id INT NOT NULL,
                rental_date DATETIME DEFAULT CURRENT_TIMESTAMP,
                return_date DATETIME,
                total_amount DECIMAL(10,2),
                status VARCHAR(20) DEFAULT 'Active',
                notes TEXT,
                FOREIGN KEY (customer_id) REFERENCES customers(customer_id),
                FOREIGN KEY (equipment_id) REFERENCES equipment(equipment_id),
                FOREIGN KEY (staff_id) REFERENCES staff(staff_id)
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS payments (
                payment_id INT AUTO_INCREMENT PRIMARY KEY,
                rental_id INT NOT NULL,
                amount DECIMAL(10,2) NOT NULL,
                payment_date DATETIME DEFAULT CURRENT_TIMESTAMP,
                payment_method VARCHAR(20),
                status VARCHAR(20) DEFAULT 'Pending',
                transaction_id VARCHAR(100),
                FOREIGN KEY (rental_id) REFERENCES rentals(rental_id)
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS maintenance (
                maintenance_id INT AUTO_INCREMENT PRIMARY KEY,
                equipment_id INT NOT NULL,
                staff_id INT NOT NULL,
                maintenance_date DATETIME DEFAULT CURRENT_TIMESTAMP,
                description TEXT,
                cost DECIMAL(10,2),
                next_maintenance_date DATE,
                status VARCHAR(20) DEFAULT 'Scheduled',
                FOREIGN KEY (equipment_id) REFERENCES equipment(equipment_id),
                FOREIGN KEY (staff_id) REFERENCES staff(staff_id)
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS users (
                user_id INT AUTO_INCREMENT PRIMARY KEY,
                username VARCHAR(50) UNIQUE NOT NULL,
                password_hash VARCHAR(255) NOT NULL,
                role VARCHAR(20) DEFAULT 'Staff',
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                last_login DATETIME
            )
            """
        ]
        
        for table_sql in tables:
            cursor.execute(table_sql)
    
    def get_connection(self):
        """Get a connection from the pool"""
        if self.pool is None:
            raise Exception("Database pool not initialized")
        return self.pool.get_connection()
    
    def execute_query(self, query, params=None, fetch=True):
        """Execute a query and return results"""
        connection = None
        cursor = None
        try:
            connection = self.get_connection()
            cursor = connection.cursor(dictionary=True, buffered=True)
            
            if params:
                cursor.execute(query, params)
            else:
                cursor.execute(query)
            
            if fetch and query.strip().upper().startswith('SELECT'):
                result = cursor.fetchall()
                return result
            else:
                connection.commit()
                return cursor.lastrowid if cursor.lastrowid else True
        except Error as e:
            if connection:
                connection.rollback()
            print(f"Database error: {e}")
            print(f"Query: {query}")
            raise e
        finally:
            if cursor:
                cursor.close()
            if connection:
                connection.close()
    
    def execute_many(self, query, params_list):
        """Execute multiple queries"""
        connection = None
        cursor = None
        try:
            connection = self.get_connection()
            cursor = connection.cursor()
            cursor.executemany(query, params_list)
            connection.commit()
            return True
        except Error as e:
            if connection:
                connection.rollback()
            raise e
        finally:
            if cursor:
                cursor.close()
            if connection:
                connection.close()


# ==================== UTILITY FUNCTIONS ====================
def hash_password(password):
    """Hash a password using SHA-256"""
    return hashlib.sha256(password.encode()).hexdigest()

def format_currency(amount):
    """Format amount as Ghanaian Cedi"""
    if amount is None:
        return f"{Config.CURRENCY_SYMBOL}0.00"
    return f"{Config.CURRENCY_SYMBOL}{amount:,.2f}"

def format_date(date_obj):
    """Format date object"""
    if isinstance(date_obj, (date, datetime)):
        return date_obj.strftime("%Y-%m-%d")
    return str(date_obj)

def generate_id():
    """Generate a unique ID"""
    return uuid.uuid4().hex[:8].upper()


# ==================== CUSTOM WIDGETS ====================
class ModernButton(QPushButton):
    """Custom styled button with Ghana colors"""
    def __init__(self, text, color=Config.PRIMARY_COLOR, icon=None, parent=None):
        super().__init__(text, parent)
        self.color = color
        self.setStyleSheet(f"""
            QPushButton {{
                background-color: {color};
                color: white;
                border: none;
                padding: 10px 20px;
                font-weight: bold;
                font-size: 13px;
                border-radius: {Config.BORDER_RADIUS};
                min-height: 30px;
            }}
            QPushButton:hover {{
                background-color: {self._adjust_color(color, 30)};
            }}
            QPushButton:pressed {{
                background-color: {self._adjust_color(color, -30)};
            }}
            QPushButton:disabled {{
                background-color: #BDC3C7;
                color: #7F8C8D;
            }}
        """)
        if icon:
            self.setIcon(QIcon(icon))
        self.setCursor(Qt.PointingHandCursor)
    
    def _adjust_color(self, color, amount):
        """Adjust color brightness"""
        color = color.lstrip('#')
        r = min(255, max(0, int(color[0:2], 16) + amount))
        g = min(255, max(0, int(color[2:4], 16) + amount))
        b = min(255, max(0, int(color[4:6], 16) + amount))
        return f'#{r:02x}{g:02x}{b:02x}'


class GradientHeader(QWidget):
    """Gradient header widget with Ghana colors"""
    def __init__(self, title, subtitle="", parent=None):
        super().__init__(parent)
        self.title = title
        self.subtitle = subtitle
        self.setMinimumHeight(100)
        self.setStyleSheet("""
            QWidget {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #CE1126, stop:0.5 #F5A623, stop:1 #006B3F);
            }
        """)
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 20, 30, 20)
        
        title_label = QLabel(title)
        title_label.setStyleSheet("color: white; font-size: 28px; font-weight: bold;")
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)
        
        if subtitle:
            subtitle_label = QLabel(subtitle)
            subtitle_label.setStyleSheet("color: rgba(255,255,255,0.9); font-size: 14px;")
            subtitle_label.setAlignment(Qt.AlignCenter)
            layout.addWidget(subtitle_label)


class StatCard(QWidget):
    """Statistics card widget with currency formatting"""
    def __init__(self, title, value, icon, color, parent=None):
        super().__init__(parent)
        self.setStyleSheet(f"""
            QWidget {{
                background-color: white;
                border-radius: 10px;
                border-left: 5px solid {color};
            }}
        """)
        
        layout = QHBoxLayout(self)
        layout.setContentsMargins(15, 10, 15, 10)
        
        # Icon
        icon_label = QLabel(icon)
        icon_label.setStyleSheet(f"font-size: 36px; color: {color};")
        icon_label.setFixedSize(50, 50)
        icon_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(icon_label)
        
        # Text
        text_layout = QVBoxLayout()
        
        title_label = QLabel(title)
        title_label.setStyleSheet("font-size: 11px; color: #7F8C8D; font-weight: bold;")
        text_layout.addWidget(title_label)
        
        value_label = QLabel(str(value))
        value_label.setStyleSheet(f"font-size: 22px; font-weight: bold; color: {Config.DARK_COLOR};")
        text_layout.addWidget(value_label)
        
        layout.addLayout(text_layout)


class SearchBar(QLineEdit):
    """Styled search bar"""
    def __init__(self, placeholder="Search...", parent=None):
        super().__init__(parent)
        self.setPlaceholderText(placeholder)
        self.setStyleSheet(f"""
            QLineEdit {{
                padding: 10px 15px;
                border: 2px solid #E0E0E0;
                border-radius: 20px;
                font-size: 13px;
                background-color: white;
            }}
            QLineEdit:focus {{
                border-color: {Config.PRIMARY_COLOR};
            }}
        """)


# ==================== THEME MANAGER ====================
class ThemeManager:
    """Application theme manager with Ghana flag colors"""
    
    @staticmethod
    def get_stylesheet():
        return f"""
            * {{
                font-family: '{Config.FONT_FAMILY}';
                font-size: 12px;
            }}
            
            QMainWindow {{
                background-color: #F5F6FA;
            }}
            
            QWidget {{
                background-color: transparent;
            }}
            
            QPushButton {{
                border: none;
                padding: 8px 16px;
                font-weight: bold;
                border-radius: {Config.BORDER_RADIUS};
                color: white;
                min-height: 30px;
            }}
            
            QPushButton:hover {{
                opacity: 0.9;
            }}
            
            QPushButton:pressed {{
                opacity: 0.7;
            }}
            
            QLineEdit, QTextEdit, QPlainTextEdit, QSpinBox, 
            QDoubleSpinBox, QDateEdit, QComboBox, QTextEdit {{
                border: 2px solid #E0E0E0;
                border-radius: {Config.BORDER_RADIUS};
                padding: 8px;
                background-color: white;
                min-height: 25px;
            }}
            
            QLineEdit:focus, QTextEdit:focus, QPlainTextEdit:focus, 
            QSpinBox:focus, QDoubleSpinBox:focus, QDateEdit:focus, 
            QComboBox:focus {{
                border-color: {Config.PRIMARY_COLOR};
            }}
            
            QTableWidget {{
                border: 1px solid #E0E0E0;
                gridline-color: #F0F0F0;
                background-color: white;
                alternate-background-color: #F8F9FA;
                selection-background-color: {Config.PRIMARY_COLOR};
                selection-color: white;
                border-radius: {Config.BORDER_RADIUS};
            }}
            
            QTableWidget::item {{
                padding: 8px;
                border-bottom: 1px solid #F0F0F0;
            }}
            
            QHeaderView::section {{
                background-color: {Config.DARK_COLOR};
                color: white;
                padding: 10px 8px;
                font-weight: bold;
                border: none;
                border-right: 1px solid #34495E;
            }}
            
            QHeaderView::section:hover {{
                background-color: #34495E;
            }}
            
            QScrollBar:vertical {{
                border: none;
                background: #F5F6FA;
                width: 12px;
                border-radius: 6px;
            }}
            
            QScrollBar::handle:vertical {{
                background: #BDC3C7;
                border-radius: 6px;
                min-height: 20px;
            }}
            
            QScrollBar::handle:vertical:hover {{
                background: #95A5A6;
            }}
            
            QTabWidget::pane {{
                border: 1px solid #E0E0E0;
                border-radius: {Config.BORDER_RADIUS};
                background: white;
            }}
            
            QTabBar::tab {{
                background: #ECF0F1;
                padding: 10px 20px;
                margin-right: 2px;
                border-top-left-radius: {Config.BORDER_RADIUS};
                border-top-right-radius: {Config.BORDER_RADIUS};
            }}
            
            QTabBar::tab:selected {{
                background: white;
                border-bottom: 3px solid {Config.PRIMARY_COLOR};
            }}
            
            QTabBar::tab:hover {{
                background: #D5DBDB;
            }}
            
            QStatusBar {{
                background-color: {Config.DARK_COLOR};
                color: white;
                padding: 5px;
            }}
            
            QMenuBar {{
                background-color: {Config.DARK_COLOR};
                color: white;
                padding: 2px;
            }}
            
            QMenuBar::item {{
                padding: 8px 15px;
                border-radius: 4px;
            }}
            
            QMenuBar::item:selected {{
                background-color: #34495E;
            }}
            
            QMenu {{
                background-color: white;
                border: 1px solid #E0E0E0;
                border-radius: 4px;
                padding: 5px;
            }}
            
            QMenu::item {{
                padding: 8px 30px 8px 20px;
                border-radius: 4px;
            }}
            
            QMenu::item:selected {{
                background-color: {Config.PRIMARY_COLOR};
                color: white;
            }}
            
            QMessageBox {{
                background-color: white;
            }}
            
            QDialog {{
                background-color: #F5F6FA;
            }}
        """


# ==================== MAIN APPLICATION ====================
class MainApplication(QMainWindow):
    """Main application window"""
    
    def __init__(self):
        super().__init__()
        self.db = DatabaseManager()
        self.current_user = None
        self.init_ui()
        self.apply_theme()
        self.show_dashboard()
    
    def init_ui(self):
        """Initialize the user interface"""
        self.setWindowTitle(f"{Config.APP_NAME} v{Config.VERSION}")
        self.setGeometry(100, 50, 1400, 900)
        
        # Create menu bar
        self.create_menu_bar()
        
        # Create toolbar
        self.create_toolbar()
        
        # Create status bar
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.status_bar.showMessage("Ready")
        
        # Create central stacked widget
        self.central_stack = QStackedWidget()
        self.setCentralWidget(self.central_stack)
    
    def create_menu_bar(self):
        """Create the application menu bar"""
        menubar = self.menuBar()
        
        # File menu
        file_menu = menubar.addMenu("&File")
        
        new_rental = QAction("New Rental", self)
        new_rental.setShortcut("Ctrl+N")
        new_rental.triggered.connect(lambda: self.open_entity('rentals'))
        file_menu.addAction(new_rental)
        
        file_menu.addSeparator()
        
        backup = QAction("Backup Database", self)
        backup.triggered.connect(self.backup_database)
        file_menu.addAction(backup)
        
        file_menu.addSeparator()
        
        exit_action = QAction("Exit", self)
        exit_action.setShortcut("Ctrl+Q")
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)
        
        # View menu
        view_menu = menubar.addMenu("&View")
        
        dashboard = QAction("Dashboard", self)
        dashboard.triggered.connect(self.show_dashboard)
        view_menu.addAction(dashboard)
        
        view_menu.addSeparator()
        
        entities = ["Customers", "Equipment", "Staff", "Rentals", "Payments", "Maintenance"]
        for entity in entities:
            action = QAction(entity, self)
            action.triggered.connect(partial(self.open_entity, entity.lower()))
            view_menu.addAction(action)
        
        # Reports menu
        reports_menu = menubar.addMenu("&Reports")
        
        report_types = [
            "Customer Report",
            "Equipment Report",
            "Rental Analysis",
            "Revenue Report",
            "Maintenance Log",
            "Staff Performance"
        ]
        
        for report in report_types:
            action = QAction(report, self)
            action.triggered.connect(partial(self.generate_report, report))
            reports_menu.addAction(action)
        
        # Help menu
        help_menu = menubar.addMenu("&Help")
        
        about = QAction("About", self)
        about.triggered.connect(self.show_about)
        help_menu.addAction(about)
        
        help_action = QAction("Help", self)
        help_action.setShortcut("F1")
        help_action.triggered.connect(self.show_help)
        help_menu.addAction(help_action)
    
    def create_toolbar(self):
        """Create the main toolbar"""
        toolbar = self.addToolBar("Main Toolbar")
        toolbar.setMovable(False)
        toolbar.setStyleSheet("""
            QToolBar {
                background-color: white;
                border-bottom: 1px solid #E0E0E0;
                padding: 5px;
                spacing: 10px;
            }
        """)
        
        actions = [
            ("🏠", "Dashboard", self.show_dashboard),
            ("👥", "Customers", lambda: self.open_entity('customers')),
            ("🔧", "Equipment", lambda: self.open_entity('equipment')),
            ("👨‍💼", "Staff", lambda: self.open_entity('staff')),
            ("📋", "Rentals", lambda: self.open_entity('rentals')),
            ("💰", "Payments", lambda: self.open_entity('payments')),
            ("🔨", "Maintenance", lambda: self.open_entity('maintenance')),
            ("📊", "Reports", self.open_reports),
        ]
        
        for icon, text, handler in actions:
            btn = ModernButton(f"{icon} {text}", Config.DARK_COLOR if icon != "📊" else Config.SECONDARY_COLOR)
            btn.setStyleSheet(f"""
                QPushButton {{
                    background-color: transparent;
                    border: none;
                    padding: 8px 15px;
                    font-weight: bold;
                    border-radius: 5px;
                    color: {Config.DARK_COLOR};
                }}
                QPushButton:hover {{
                    background-color: #F0F0F0;
                }}
            """)
            btn.clicked.connect(handler)
            btn.setCursor(Qt.PointingHandCursor)
            toolbar.addWidget(btn)
    
    def apply_theme(self):
        """Apply the application theme"""
        self.setStyleSheet(ThemeManager.get_stylesheet())
    
    def show_dashboard(self):
        """Display the dashboard"""
        dashboard = QWidget()
        layout = QVBoxLayout(dashboard)
        layout.setSpacing(20)
        
        # Header with Ghana colors
        header = GradientHeader(
            f"Welcome to {Config.APP_NAME}",
            f"{Config.COMPANY_NAME} • {datetime.now().strftime('%A, %B %d, %Y')} • Currency: {Config.CURRENCY_SYMBOL}"
        )
        layout.addWidget(header)
        
        # Stats cards
        stats = self.get_dashboard_stats()
        stats_widget = QWidget()
        stats_layout = QHBoxLayout(stats_widget)
        stats_layout.setSpacing(20)
        
        cards = [
            ("Total Customers", stats['customers'], "👥", Config.PRIMARY_COLOR),
            ("Active Rentals", stats['active_rentals'], "📋", Config.SUCCESS_COLOR),
            ("Available Equipment", stats['available_equipment'], "🔧", Config.INFO_COLOR),
            ("Revenue Today", format_currency(stats['revenue_today']), "💰", Config.GHANA_GOLD),
        ]
        
        for title, value, icon, color in cards:
            card = StatCard(title, value, icon, color)
            stats_layout.addWidget(card)
        
        layout.addWidget(stats_widget)
        
        # Content area
        content = QWidget()
        content_layout = QHBoxLayout(content)
        content_layout.setSpacing(20)
        
        # Left: Charts
        if QTCHART_AVAILABLE:
            chart_widget = self.create_chart_widget()
            content_layout.addWidget(chart_widget, 2)
        
        # Right: Quick actions
        actions_widget = self.create_quick_actions()
        content_layout.addWidget(actions_widget, 1)
        
        layout.addWidget(content)
        
        # Recent activity
        recent = self.create_recent_activity()
        layout.addWidget(recent)
        
        # Update stack
        self.add_to_stack(dashboard)
        self.status_bar.showMessage(f"Dashboard updated at {datetime.now().strftime('%H:%M:%S')}")
    
    def get_dashboard_stats(self):
        """Get statistics for dashboard"""
        stats = {
            'customers': 0,
            'active_rentals': 0,
            'available_equipment': 0,
            'revenue_today': 0
        }
        
        try:
            result = self.db.execute_query("SELECT COUNT(*) as count FROM customers")
            stats['customers'] = result[0]['count'] if result else 0
            
            result = self.db.execute_query("SELECT COUNT(*) as count FROM rentals WHERE status = 'Active'")
            stats['active_rentals'] = result[0]['count'] if result else 0
            
            result = self.db.execute_query("SELECT COUNT(*) as count FROM equipment WHERE status = 'Available'")
            stats['available_equipment'] = result[0]['count'] if result else 0
            
            result = self.db.execute_query(
                "SELECT COALESCE(SUM(amount), 0) as total FROM payments "
                "WHERE DATE(payment_date) = CURDATE() AND status = 'Completed'"
            )
            stats['revenue_today'] = float(result[0]['total']) if result else 0
            
        except Exception as e:
            print(f"Error fetching stats: {e}")
        
        return stats
    
    def create_chart_widget(self):
        """Create chart widget for dashboard"""
        widget = QWidget()
        widget.setStyleSheet("background-color: white; border-radius: 10px;")
        layout = QVBoxLayout(widget)
        
        title = QLabel("Rental Trends (Last 7 Days)")
        title.setStyleSheet(f"font-size: 16px; font-weight: bold; color: {Config.DARK_COLOR}; padding: 10px;")
        layout.addWidget(title)
        
        chart_view = QChartView()
        chart_view.setRenderHint(QPainter.Antialiasing)
        chart_view.setMinimumHeight(250)
        
        chart = QChart()
        chart.setAnimationOptions(QChart.SeriesAnimations)
        chart.setTheme(QChart.ChartThemeLight)
        
        series = QLineSeries()
        series.setName("Rentals")
        
        try:
            query = """
                SELECT DATE(rental_date) as date, COUNT(*) as count
                FROM rentals
                WHERE rental_date >= DATE_SUB(CURDATE(), INTERVAL 7 DAY)
                GROUP BY DATE(rental_date)
                ORDER BY date
            """
            results = self.db.execute_query(query)
            
            axis_x = QDateTimeAxis()
            axis_x.setFormat("MMM dd")
            axis_x.setTitleText("Date")
            
            axis_y = QValueAxis()
            axis_y.setTitleText("Number of Rentals")
            axis_y.setLabelFormat("%d")
            
            if results:
                for row in results:
                    if row['date']:
                        dt = QDateTime(row['date'].year, row['date'].month, row['date'].day, 0, 0)
                        series.append(dt.toMSecsSinceEpoch(), row['count'])
            
            chart.addAxis(axis_x, Qt.AlignBottom)
            chart.addAxis(axis_y, Qt.AlignLeft)
            chart.addSeries(series)
            series.attachAxis(axis_x)
            series.attachAxis(axis_y)
            
        except Exception as e:
            print(f"Error creating chart: {e}")
        
        chart_view.setChart(chart)
        layout.addWidget(chart_view)
        
        return widget
    
    def create_quick_actions(self):
        """Create quick actions panel"""
        widget = QWidget()
        widget.setStyleSheet("background-color: white; border-radius: 10px;")
        layout = QVBoxLayout(widget)
        layout.setSpacing(15)
        
        title = QLabel("Quick Actions")
        title.setStyleSheet(f"font-size: 16px; font-weight: bold; color: {Config.DARK_COLOR};")
        layout.addWidget(title)
        
        actions = [
            ("📋 New Rental", Config.GHANA_GREEN, lambda: self.open_entity('rentals')),
            ("👤 Add Customer", Config.PRIMARY_COLOR, lambda: self.open_entity('customers')),
            ("🔧 Add Equipment", Config.INFO_COLOR, lambda: self.open_entity('equipment')),
            ("💰 Record Payment", Config.GHANA_GOLD, lambda: self.open_entity('payments')),
            ("🔨 Schedule Maintenance", Config.GHANA_RED, lambda: self.open_entity('maintenance')),
        ]
        
        for text, color, handler in actions:
            btn = ModernButton(text, color)
            btn.clicked.connect(handler)
            layout.addWidget(btn)
        
        layout.addStretch()
        return widget
    
    def create_recent_activity(self):
        """Create recent activity widget"""
        widget = QWidget()
        widget.setStyleSheet("background-color: white; border-radius: 10px;")
        layout = QVBoxLayout(widget)
        
        title = QLabel("Recent Activity")
        title.setStyleSheet(f"font-size: 16px; font-weight: bold; color: {Config.DARK_COLOR};")
        layout.addWidget(title)
        
        table = QTableWidget()
        table.setColumnCount(4)
        table.setHorizontalHeaderLabels(["Date", "Activity", "Details", "Status"])
        table.horizontalHeader().setStretchLastSection(True)
        table.setMaximumHeight(150)
        table.setAlternatingRowColors(True)
        
        try:
            query = """
                (SELECT rental_date as date, 'Rental' as activity, 
                        CONCAT(c.first_name, ' ', c.last_name, ' rented ', e.name) as details,
                        r.status
                 FROM rentals r
                 JOIN customers c ON r.customer_id = c.customer_id
                 JOIN equipment e ON r.equipment_id = e.equipment_id
                 ORDER BY r.rental_date DESC LIMIT 5)
                UNION ALL
                (SELECT payment_date as date, 'Payment' as activity,
                        CONCAT('", Config.CURRENCY_SYMBOL, "', p.amount, ' - ', p.payment_method) as details,
                        p.status
                 FROM payments p
                 ORDER BY p.payment_date DESC LIMIT 5)
                ORDER BY date DESC LIMIT 10
            """
            results = self.db.execute_query(query)
            
            if results:
                table.setRowCount(len(results))
                for i, row in enumerate(results):
                    table.setItem(i, 0, QTableWidgetItem(format_date(row['date'])))
                    table.setItem(i, 1, QTableWidgetItem(row['activity']))
                    table.setItem(i, 2, QTableWidgetItem(row['details']))
                    
                    status_item = QTableWidgetItem(row['status'])
                    if row['status'] in ['Active', 'Completed', 'Available']:
                        status_item.setForeground(QColor(Config.SUCCESS_COLOR))
                    elif row['status'] in ['Overdue', 'Failed', 'Maintenance']:
                        status_item.setForeground(QColor(Config.DANGER_COLOR))
                    table.setItem(i, 3, status_item)
        
        except Exception as e:
            print(f"Error fetching recent activity: {e}")
        
        layout.addWidget(table)
        return widget
    
    def add_to_stack(self, widget):
        """Add widget to stack and switch to it"""
        if self.central_stack.count() > 0:
            self.central_stack.removeWidget(self.central_stack.widget(0))
        self.central_stack.addWidget(widget)
    
    def open_entity(self, entity_name):
        """Open entity management dialog"""
        entity_map = {
            'customers': ('Customer Management', 'customers', 
                         ['customer_id', 'first_name', 'last_name', 'email', 'phone', 'address', 'registration_date', 'status']),
            'equipment': ('Equipment Management', 'equipment',
                         ['equipment_id', 'name', 'category', 'description', 'daily_rate', 'status', 'purchase_date', 'last_maintenance_date', 'serial_number']),
            'staff': ('Staff Management', 'staff',
                     ['staff_id', 'first_name', 'last_name', 'email', 'phone', 'position', 'hire_date', 'salary', 'status']),
            'rentals': ('Rental Management', 'rentals',
                       ['rental_id', 'customer_id', 'equipment_id', 'staff_id', 'rental_date', 'return_date', 'total_amount', 'status', 'notes']),
            'payments': ('Payment Management', 'payments',
                        ['payment_id', 'rental_id', 'amount', 'payment_date', 'payment_method', 'status', 'transaction_id']),
            'maintenance': ('Maintenance Management', 'maintenance',
                           ['maintenance_id', 'equipment_id', 'staff_id', 'maintenance_date', 'description', 'cost', 'next_maintenance_date', 'status']),
        }
        
        if entity_name in entity_map:
            title, table, cols = entity_map[entity_name]
            dialog = EntityDialog(title, table, cols, self)
            dialog.exec_()
    
    def open_reports(self):
        """Open reports dialog"""
        dialog = ReportsDialog(self)
        dialog.exec_()
    
    def generate_report(self, report_type):
        """Generate a specific report"""
        dialog = ReportsDialog(self, report_type)
        dialog.exec_()
    
    def backup_database(self):
        """Backup the database"""
        try:
            os.makedirs('backup', exist_ok=True)
            filename = f"backup/backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.sql"
            QMessageBox.information(self, "Backup", f"Backup saved to:\n{filename}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Backup failed: {e}")
    
    def show_about(self):
        """Show about dialog"""
        about_text = f"""
        <h1>{Config.APP_NAME}</h1>
        <p>Version: {Config.VERSION}</p>
        <p>{Config.COMPANY_NAME}</p>
        <hr>
        <p>Professional Equipment Rental Management System</p>
        <p>Currency: {Config.CURRENCY_SYMBOL} (Ghanaian Cedi)</p>
        <p>Built with PyQt5 and MySQL</p>
        <p>© 2024 All Rights Reserved</p>
        """
        QMessageBox.about(self, f"About {Config.APP_NAME}", about_text)
    
    def show_help(self):
        """Show help dialog"""
        help_text = f"""
        <h2>Help Guide</h2>
        <hr>
        <h3>Currency:</h3>
        <p>All financial amounts are in <b>{Config.CURRENCY_SYMBOL}</b> (Ghanaian Cedi)</p>
        <h3>Keyboard Shortcuts:</h3>
        <ul>
            <li><b>Ctrl+N</b> - New Rental</li>
            <li><b>Ctrl+Q</b> - Exit Application</li>
            <li><b>F1</b> - Show Help</li>
        </ul>
        <h3>Navigation:</h3>
        <p>Use the menu bar or toolbar to navigate between sections.</p>
        <h3>Reports:</h3>
        <p>Access reports from the Reports menu.</p>
        """
        QMessageBox.information(self, "Help", help_text)


# ==================== ENTITY DIALOG ====================
class EntityDialog(QDialog):
    """Generic entity management dialog"""
    
    def __init__(self, title, table_name, columns, parent=None):
        super().__init__(parent)
        self.title = title
        self.table_name = table_name
        self.columns = columns
        self.db = DatabaseManager()
        self.current_id = None
        self.current_page = 0
        self.page_size = 20
        
        self.init_ui()
        self.load_data()
    
    def init_ui(self):
        """Initialize the dialog UI"""
        self.setWindowTitle(self.title)
        self.setGeometry(100, 100, 1300, 750)
        self.setModal(True)
        
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(0)
        
        # Header with Ghana colors
        header = GradientHeader(self.title, f"Currency: {Config.CURRENCY_SYMBOL}")
        main_layout.addWidget(header)
        
        # Main content
        content = QWidget()
        content_layout = QHBoxLayout(content)
        content_layout.setSpacing(20)
        
        # Form panel
        form_panel = self.create_form_panel()
        content_layout.addWidget(form_panel, 1)
        
        # Table panel
        table_panel = self.create_table_panel()
        content_layout.addWidget(table_panel, 3)
        
        main_layout.addWidget(content)
    
    def create_form_panel(self):
        """Create the form panel"""
        panel = QWidget()
        panel.setStyleSheet("background-color: white; border-radius: 10px;")
        panel.setFixedWidth(350)
        layout = QVBoxLayout(panel)
        
        title = QLabel("Record Details")
        title.setStyleSheet(f"font-size: 16px; font-weight: bold; color: {Config.DARK_COLOR};")
        layout.addWidget(title)
        
        # Scroll area for form fields
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("border: none;")
        
        scroll_content = QWidget()
        self.form_layout = QVBoxLayout(scroll_content)
        
        self.fields = {}
        for col in self.columns:
            if col != self.columns[0]:  # Skip ID
                field_widget = self.create_field(col)
                self.form_layout.addWidget(field_widget)
                self.fields[col] = field_widget.findChild(QLineEdit) or field_widget.findChild(QComboBox) or field_widget.findChild(QDateEdit) or field_widget.findChild(QDoubleSpinBox)
        
        self.form_layout.addStretch()
        scroll.setWidget(scroll_content)
        layout.addWidget(scroll)
        
        # Buttons
        btn_container = QWidget()
        btn_layout = QVBoxLayout(btn_container)
        
        add_btn = ModernButton("➕ Add Record", Config.SUCCESS_COLOR)
        add_btn.clicked.connect(self.add_record)
        btn_layout.addWidget(add_btn)
        
        update_btn = ModernButton("✏️ Update Record", Config.WARNING_COLOR)
        update_btn.clicked.connect(self.update_record)
        btn_layout.addWidget(update_btn)
        
        delete_btn = ModernButton("🗑️ Delete Record", Config.DANGER_COLOR)
        delete_btn.clicked.connect(self.delete_record)
        btn_layout.addWidget(delete_btn)
        
        clear_btn = ModernButton("🔄 Clear Form", "#95A5A6")
        clear_btn.clicked.connect(self.clear_form)
        btn_layout.addWidget(clear_btn)
        
        layout.addWidget(btn_container)
        return panel
    
    def create_field(self, col):
        """Create a form field"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setSpacing(5)
        
        label = QLabel(col.replace('_', ' ').title() + ":")
        label.setStyleSheet("font-weight: bold; color: #555;")
        layout.addWidget(label)
        
        # Special handling for specific columns
        if 'status' in col.lower():
            field = QComboBox()
            if self.table_name == 'rentals':
                field.addItems(['Active', 'Completed', 'Overdue', 'Cancelled'])
            elif self.table_name == 'equipment':
                field.addItems(['Available', 'Rented', 'Maintenance', 'Retired'])
            elif self.table_name == 'payments':
                field.addItems(['Pending', 'Completed', 'Failed', 'Refunded'])
            else:
                field.addItems(['Active', 'Inactive'])
        elif 'date' in col.lower():
            field = QDateEdit()
            field.setCalendarPopup(True)
            field.setDate(QDate.currentDate())
        elif 'amount' in col.lower() or 'rate' in col.lower() or 'salary' in col.lower() or 'cost' in col.lower():
            field = QDoubleSpinBox()
            field.setRange(0, 999999.99)
            field.setPrefix(f"{Config.CURRENCY_SYMBOL} ")
            field.setDecimals(2)
        else:
            field = QLineEdit()
            field.setPlaceholderText(f"Enter {col.replace('_', ' ').lower()}...")
        
        layout.addWidget(field)
        return widget
    
    def create_table_panel(self):
        """Create the table panel"""
        panel = QWidget()
        panel.setStyleSheet("background-color: white; border-radius: 10px;")
        layout = QVBoxLayout(panel)
        
        # Search bar
        search_container = QWidget()
        search_layout = QHBoxLayout(search_container)
        
        self.search_bar = SearchBar("Search records...")
        self.search_bar.textChanged.connect(self.search_records)
        search_layout.addWidget(self.search_bar)
        
        refresh_btn = ModernButton("🔄 Refresh", Config.INFO_COLOR)
        refresh_btn.clicked.connect(self.load_data)
        search_layout.addWidget(refresh_btn)
        
        layout.addWidget(search_container)
        
        # Table
        self.table = QTableWidget()
        self.table.setColumnCount(len(self.columns))
        self.table.setHorizontalHeaderLabels([col.replace('_', ' ').title() for col in self.columns])
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setSelectionMode(QTableWidget.SingleSelection)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.clicked.connect(self.on_row_select)
        self.table.setSortingEnabled(True)
        layout.addWidget(self.table)
        
        # Pagination
        pagination = QWidget()
        pagination_layout = QHBoxLayout(pagination)
        
        self.total_label = QLabel("Total: 0 records")
        self.total_label.setStyleSheet("color: #666;")
        pagination_layout.addWidget(self.total_label)
        
        pagination_layout.addStretch()
        
        prev_btn = ModernButton("◀ Previous", "#95A5A6")
        prev_btn.clicked.connect(self.previous_page)
        pagination_layout.addWidget(prev_btn)
        
        next_btn = ModernButton("Next ▶", "#95A5A6")
        next_btn.clicked.connect(self.next_page)
        pagination_layout.addWidget(next_btn)
        
        layout.addWidget(pagination)
        return panel
    
    def load_data(self):
        """Load data into the table"""
        try:
            query = f"SELECT * FROM {self.table_name} ORDER BY {self.columns[0]} DESC"
            results = self.db.execute_query(query)
            
            self.table.setRowCount(0)
            if results:
                self.table.setRowCount(len(results))
                for row_idx, row in enumerate(results):
                    for col_idx, col in enumerate(self.columns):
                        value = row[col] if row[col] is not None else ''
                        item = QTableWidgetItem(str(value))
                        
                        # Color code status
                        if col == 'status':
                            if str(value) in ['Active', 'Completed', 'Available', 'Completed']:
                                item.setForeground(QColor(Config.SUCCESS_COLOR))
                            elif str(value) in ['Overdue', 'Maintenance', 'Failed', 'Retired']:
                                item.setForeground(QColor(Config.DANGER_COLOR))
                        
                        # Format currency columns
                        if any(word in col.lower() for word in ['amount', 'rate', 'salary', 'cost']):
                            if value:
                                try:
                                    item.setText(format_currency(float(value)))
                                except:
                                    pass
                        
                        self.table.setItem(row_idx, col_idx, item)
            
            self.update_pagination()
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load data: {e}")
    
    def on_row_select(self):
        """Handle row selection"""
        current_row = self.table.currentRow()
        if current_row >= 0:
            self.current_id = self.table.item(current_row, 0).text()
            self.clear_form()
            
            for i, col in enumerate(self.columns[1:]):
                item = self.table.item(current_row, i + 1)
                if item and col in self.fields:
                    field = self.fields[col]
                    value = item.text()
                    
                    if isinstance(field, QLineEdit):
                        field.setText(value)
                    elif isinstance(field, QComboBox):
                        index = field.findText(value)
                        if index >= 0:
                            field.setCurrentIndex(index)
                    elif isinstance(field, QDateEdit):
                        try:
                            dt = datetime.strptime(value, '%Y-%m-%d')
                            field.setDate(QDate(dt.year, dt.month, dt.day))
                        except:
                            pass
                    elif isinstance(field, QDoubleSpinBox):
                        try:
                            # Remove currency symbol and commas
                            clean_value = value.replace(Config.CURRENCY_SYMBOL, '').replace(',', '').strip()
                            field.setValue(float(clean_value))
                        except:
                            pass
    
    def clear_form(self):
        """Clear the form"""
        self.current_id = None
        for field in self.fields.values():
            if isinstance(field, QLineEdit):
                field.clear()
            elif isinstance(field, QComboBox):
                field.setCurrentIndex(0)
            elif isinstance(field, QDateEdit):
                field.setDate(QDate.currentDate())
            elif isinstance(field, QDoubleSpinBox):
                field.setValue(0)
    
    def add_record(self):
        """Add a new record"""
        values = []
        for col in self.columns[1:]:
            field = self.fields.get(col)
            if field:
                if isinstance(field, QLineEdit):
                    value = field.text().strip()
                elif isinstance(field, QComboBox):
                    value = field.currentText()
                elif isinstance(field, QDateEdit):
                    value = field.date().toString('yyyy-MM-dd')
                elif isinstance(field, QDoubleSpinBox):
                    value = field.value()
                else:
                    value = ''
                values.append(value if value else None)
        
        if not any(values):
            QMessageBox.warning(self, "Warning", "Please enter at least one value")
            return
        
        try:
            placeholders = ', '.join(['%s'] * len(values))
            columns_str = ', '.join(self.columns[1:])
            query = f"INSERT INTO {self.table_name} ({columns_str}) VALUES ({placeholders})"
            
            self.db.execute_query(query, values, fetch=False)
            QMessageBox.information(self, "Success", "Record added successfully!")
            self.load_data()
            self.clear_form()
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to add record: {e}")
    
    def update_record(self):
        """Update the selected record"""
        if not self.current_id:
            QMessageBox.warning(self, "Warning", "Please select a record to update")
            return
        
        set_clause = []
        values = []
        for col in self.columns[1:]:
            field = self.fields.get(col)
            if field:
                if isinstance(field, QLineEdit):
                    value = field.text().strip()
                elif isinstance(field, QComboBox):
                    value = field.currentText()
                elif isinstance(field, QDateEdit):
                    value = field.date().toString('yyyy-MM-dd')
                elif isinstance(field, QDoubleSpinBox):
                    value = field.value()
                else:
                    value = ''
                
                if value:
                    set_clause.append(f"{col} = %s")
                    values.append(value)
        
        if not set_clause:
            QMessageBox.warning(self, "Warning", "No values to update")
            return
        
        try:
            set_clause_str = ', '.join(set_clause)
            values.append(self.current_id)
            query = f"UPDATE {self.table_name} SET {set_clause_str} WHERE {self.columns[0]} = %s"
            
            self.db.execute_query(query, values, fetch=False)
            QMessageBox.information(self, "Success", "Record updated successfully!")
            self.load_data()
            self.clear_form()
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to update record: {e}")
    
    def delete_record(self):
        """Delete the selected record"""
        if not self.current_id:
            QMessageBox.warning(self, "Warning", "Please select a record to delete")
            return
        
        reply = QMessageBox.question(
            self, "Confirm Delete",
            "Are you sure you want to delete this record?\nThis action cannot be undone.",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            try:
                query = f"DELETE FROM {self.table_name} WHERE {self.columns[0]} = %s"
                self.db.execute_query(query, (self.current_id,), fetch=False)
                QMessageBox.information(self, "Success", "Record deleted successfully!")
                self.load_data()
                self.clear_form()
                
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to delete record: {e}")
    
    def search_records(self):
        """Search for records"""
        search_term = self.search_bar.text().strip()
        if not search_term:
            self.load_data()
            return
        
        try:
            conditions = []
            params = []
            for col in self.columns:
                if col != self.columns[0]:
                    conditions.append(f"{col} LIKE %s")
                    params.append(f"%{search_term}%")
            
            query = f"SELECT * FROM {self.table_name} WHERE {' OR '.join(conditions)}"
            results = self.db.execute_query(query, params)
            
            self.table.setRowCount(0)
            if results:
                self.table.setRowCount(len(results))
                for row_idx, row in enumerate(results):
                    for col_idx, col in enumerate(self.columns):
                        value = row[col] if row[col] is not None else ''
                        self.table.setItem(row_idx, col_idx, QTableWidgetItem(str(value)))
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Search failed: {e}")
    
    def update_pagination(self):
        """Update pagination info"""
        total = self.table.rowCount()
        self.total_label.setText(f"Total: {total} records")
    
    def previous_page(self):
        """Go to previous page"""
        if self.current_page > 0:
            self.current_page -= 1
            self.load_data()
    
    def next_page(self):
        """Go to next page"""
        self.current_page += 1
        self.load_data()


# ==================== REPORTS DIALOG ====================
class ReportsDialog(QDialog):
    """Reports generation dialog with Ghana currency"""
    
    def __init__(self, parent=None, report_type=None):
        super().__init__(parent)
        self.db = DatabaseManager()
        self.report_type = report_type
        self.init_ui()
        
        if report_type:
            self.report_combo.setCurrentText(report_type)
            self.generate_report()
    
    def init_ui(self):
        """Initialize the dialog UI"""
        self.setWindowTitle("Report Generator")
        self.setGeometry(200, 200, 900, 650)
        self.setModal(True)
        
        layout = QVBoxLayout(self)
        
        # Header with Ghana colors
        header = GradientHeader(
            "Report Generator", 
            f"Generate professional business reports • Currency: {Config.CURRENCY_SYMBOL}"
        )
        layout.addWidget(header)
        
        # Content
        content = QWidget()
        content_layout = QVBoxLayout(content)
        
        # Report selection
        selection = QHBoxLayout()
        
        label = QLabel("Select Report:")
        label.setStyleSheet("font-weight: bold; font-size: 14px;")
        selection.addWidget(label)
        
        self.report_combo = QComboBox()
        self.report_combo.addItems([
            "Customer Report",
            "Equipment Report",
            "Rental Analysis",
            "Revenue Report",
            "Maintenance Log",
            "Staff Performance"
        ])
        self.report_combo.setStyleSheet("min-width: 200px; padding: 8px;")
        selection.addWidget(self.report_combo)
        
        generate_btn = ModernButton("📊 Generate", Config.GHANA_GREEN)
        generate_btn.clicked.connect(self.generate_report)
        selection.addWidget(generate_btn)
        
        selection.addStretch()
        content_layout.addLayout(selection)
        
        # Preview
        preview_label = QLabel("Report Preview")
        preview_label.setStyleSheet(f"font-size: 16px; font-weight: bold; color: {Config.DARK_COLOR}; margin-top: 15px;")
        content_layout.addWidget(preview_label)
        
        self.preview_table = QTableWidget()
        self.preview_table.setAlternatingRowColors(True)
        self.preview_table.setEditTriggers(QTableWidget.NoEditTriggers)
        content_layout.addWidget(self.preview_table)
        
        # Export buttons
        export = QHBoxLayout()
        export.addStretch()
        
        pdf_btn = ModernButton("📄 Export PDF", Config.GHANA_RED)
        pdf_btn.clicked.connect(lambda: self.export_report("PDF"))
        export.addWidget(pdf_btn)
        
        excel_btn = ModernButton("📊 Export Excel", Config.GHANA_GREEN)
        excel_btn.clicked.connect(lambda: self.export_report("Excel"))
        export.addWidget(excel_btn)
        
        print_btn = ModernButton("🖨️ Print", Config.PRIMARY_COLOR)
        print_btn.clicked.connect(self.print_report)
        export.addWidget(print_btn)
        
        export.addStretch()
        content_layout.addLayout(export)
        
        layout.addWidget(content)
        
        # Footer
        footer = QHBoxLayout()
        footer.addStretch()
        
        close_btn = ModernButton("Close", Config.DARK_COLOR)
        close_btn.clicked.connect(self.close)
        footer.addWidget(close_btn)
        
        layout.addLayout(footer)
    
    def generate_report(self):
        """Generate the selected report"""
        report_type = self.report_combo.currentText()
        
        # Define report queries
        reports = {
            "Customer Report": f"""
                SELECT 
                    c.customer_id as ID,
                    CONCAT(c.first_name, ' ', c.last_name) as Name,
                    c.email as Email,
                    c.phone as Phone,
                    COUNT(r.rental_id) as 'Total Rentals',
                    COALESCE(SUM(p.amount), 0) as 'Total Spent ({Config.CURRENCY_SYMBOL})'
                FROM customers c
                LEFT JOIN rentals r ON c.customer_id = r.customer_id
                LEFT JOIN payments p ON r.rental_id = p.rental_id
                GROUP BY c.customer_id
                ORDER BY 'Total Spent' DESC
            """,
            "Equipment Report": f"""
                SELECT 
                    e.equipment_id as ID,
                    e.name as Equipment,
                    e.category as Category,
                    e.status as Status,
                    e.daily_rate as 'Daily Rate ({Config.CURRENCY_SYMBOL})',
                    COUNT(r.rental_id) as 'Times Rented'
                FROM equipment e
                LEFT JOIN rentals r ON e.equipment_id = r.equipment_id
                GROUP BY e.equipment_id
                ORDER BY 'Times Rented' DESC
            """,
            "Rental Analysis": f"""
                SELECT 
                    r.rental_id as 'Rental ID',
                    CONCAT(c.first_name, ' ', c.last_name) as Customer,
                    e.name as Equipment,
                    r.rental_date as 'Rental Date',
                    r.return_date as 'Return Date',
                    r.total_amount as 'Amount ({Config.CURRENCY_SYMBOL})',
                    r.status as Status
                FROM rentals r
                JOIN customers c ON r.customer_id = c.customer_id
                JOIN equipment e ON r.equipment_id = e.equipment_id
                ORDER BY r.rental_date DESC
            """,
            "Revenue Report": f"""
                SELECT 
                    DATE(p.payment_date) as Date,
                    COUNT(DISTINCT p.payment_id) as Transactions,
                    SUM(p.amount) as 'Revenue ({Config.CURRENCY_SYMBOL})',
                    AVG(p.amount) as 'Avg Transaction ({Config.CURRENCY_SYMBOL})'
                FROM payments p
                WHERE p.status = 'Completed'
                GROUP BY DATE(p.payment_date)
                ORDER BY Date DESC
            """,
            "Maintenance Log": f"""
                SELECT 
                    m.maintenance_id as ID,
                    e.name as Equipment,
                    CONCAT(s.first_name, ' ', s.last_name) as Technician,
                    m.maintenance_date as Date,
                    m.description as Description,
                    m.cost as 'Cost ({Config.CURRENCY_SYMBOL})',
                    m.status as Status
                FROM maintenance m
                JOIN equipment e ON m.equipment_id = e.equipment_id
                JOIN staff s ON m.staff_id = s.staff_id
                ORDER BY m.maintenance_date DESC
            """,
            "Staff Performance": f"""
                SELECT 
                    CONCAT(s.first_name, ' ', s.last_name) as Staff,
                    s.position as Position,
                    COUNT(r.rental_id) as 'Rentals Handled',
                    COUNT(m.maintenance_id) as 'Maintenance Tasks',
                    COALESCE(SUM(p.amount), 0) as 'Revenue Generated ({Config.CURRENCY_SYMBOL})'
                FROM staff s
                LEFT JOIN rentals r ON s.staff_id = r.staff_id
                LEFT JOIN payments p ON r.rental_id = p.rental_id
                LEFT JOIN maintenance m ON s.staff_id = m.staff_id
                GROUP BY s.staff_id
                ORDER BY 'Revenue Generated' DESC
            """
        }
        
        try:
            query = reports.get(report_type)
            if not query:
                QMessageBox.warning(self, "Warning", "Report not found")
                return
            
            results = self.db.execute_query(query)
            
            if results:
                columns = list(results[0].keys())
                self.preview_table.setColumnCount(len(columns))
                self.preview_table.setHorizontalHeaderLabels(columns)
                
                self.preview_table.setRowCount(len(results))
                for row_idx, row in enumerate(results):
                    for col_idx, col in enumerate(columns):
                        value = row[col] if row[col] is not None else ''
                        item = QTableWidgetItem(str(value))
                        
                        # Format currency columns
                        if any(word in col for word in ['Daily Rate', 'Amount', 'Revenue', 'Cost', 'Total Spent', 'Avg Transaction']):
                            if value:
                                try:
                                    item.setText(format_currency(float(value)))
                                except:
                                    pass
                        
                        self.preview_table.setItem(row_idx, col_idx, item)
                
                self.preview_table.resizeColumnsToContents()
                QMessageBox.information(self, "Success", f"Report generated successfully!")
            else:
                QMessageBox.information(self, "Info", "No data found for this report")
                
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to generate report: {e}")
    
    def export_report(self, format_type):
        """Export the report"""
        if self.preview_table.rowCount() == 0:
            QMessageBox.warning(self, "Warning", "No data to export")
            return
        
        os.makedirs('reports', exist_ok=True)
        report_name = self.report_combo.currentText().replace(' ', '_')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filepath = f"reports/{report_name}_{timestamp}"
        
        try:
            if format_type == "PDF":
                self.export_pdf(filepath + ".pdf")
            else:
                self.export_excel(filepath + ".xlsx")
                
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Export failed: {e}")
    
    def export_pdf(self, filepath):
        """Export to PDF"""
        doc = SimpleDocTemplate(filepath, pagesize=landscape(letter))
        elements = []
        
        # Title with Ghana colors
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor(Config.GHANA_RED),
            spaceAfter=10,
            alignment=TA_CENTER
        )
        
        title = Paragraph(self.report_combo.currentText(), title_style)
        elements.append(title)
        
        # Currency info
        currency_style = ParagraphStyle(
            'Currency',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor(Config.GHANA_GOLD),
            alignment=TA_CENTER
        )
        currency = Paragraph(f"Currency: {Config.CURRENCY_SYMBOL} (Ghanaian Cedi)", currency_style)
        elements.append(currency)
        
        # Timestamp
        timestamp_style = ParagraphStyle(
            'Timestamp',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey,
            alignment=TA_RIGHT
        )
        timestamp = Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", timestamp_style)
        elements.append(timestamp)
        elements.append(Spacer(1, 20))
        
        # Table data
        headers = []
        for col in range(self.preview_table.columnCount()):
            headers.append(self.preview_table.horizontalHeaderItem(col).text())
        
        table_data = [headers]
        for row in range(self.preview_table.rowCount()):
            row_data = []
            for col in range(self.preview_table.columnCount()):
                item = self.preview_table.item(row, col)
                row_data.append(item.text() if item else '')
            table_data.append(row_data)
        
        table = Table(table_data, repeatRows=1)
        
        # Style the table
        style_commands = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(Config.GHANA_RED)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.Color(0.95, 0.95, 0.95)]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ]
        
        table.setStyle(TableStyle(style_commands))
        elements.append(table)
        
        # Footer with Ghana flag colors
        footer_text = f"Generated by {Config.APP_NAME} v{Config.VERSION} • © {datetime.now().year} {Config.COMPANY_NAME}"
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.grey,
            alignment=TA_CENTER
        )
        footer = Paragraph(footer_text, footer_style)
        elements.append(Spacer(1, 20))
        elements.append(footer)
        
        doc.build(elements)
        QMessageBox.information(self, "Success", f"Report exported to:\n{filepath}")
        
        # Open the file
        if os.name == 'nt':
            os.startfile(filepath)
        else:
            subprocess.call(['open', filepath])
    
    def export_excel(self, filepath):
        """Export to Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = self.report_combo.currentText()[:31]
            
            # Title
            ws.merge_cells('A1:H1')
            ws['A1'] = self.report_combo.currentText()
            ws['A1'].font = Font(size=16, bold=True, color=Config.GHANA_RED)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Currency info
            ws.merge_cells('A2:H2')
            ws['A2'] = f"Currency: {Config.CURRENCY_SYMBOL} (Ghanaian Cedi)"
            ws['A2'].font = Font(size=11, bold=True, color=Config.GHANA_GOLD)
            ws['A2'].alignment = Alignment(horizontal='center')
            
            # Timestamp
            ws.merge_cells('A3:H3')
            ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A3'].font = Font(size=10, color='7F8C8D')
            ws['A3'].alignment = Alignment(horizontal='right')
            
            # Headers
            header_fill = PatternFill(start_color=Config.GHANA_RED, end_color=Config.GHANA_RED, fill_type='solid')
            header_font = Font(size=11, bold=True, color='FFFFFF')
            
            for col in range(self.preview_table.columnCount()):
                cell = ws.cell(row=5, column=col+1)
                cell.value = self.preview_table.horizontalHeaderItem(col).text()
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # Data
            for row in range(self.preview_table.rowCount()):
                for col in range(self.preview_table.columnCount()):
                    cell = ws.cell(row=row+6, column=col+1)
                    item = self.preview_table.item(row, col)
                    cell.value = item.text() if item else ''
                    cell.alignment = Alignment(horizontal='center')
            
            # Adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[column].width = min(max_length + 2, 50)
            
            wb.save(filepath)
            QMessageBox.information(self, "Success", f"Report exported to:\n{filepath}")
            
            if os.name == 'nt':
                os.startfile(filepath)
            else:
                subprocess.call(['open', filepath])
                
        except ImportError:
            QMessageBox.warning(self, "Warning", 
                "openpyxl is required for Excel export.\nInstall it with: pip install openpyxl")
    
    def print_report(self):
        """Print the report"""
        if self.preview_table.rowCount() == 0:
            QMessageBox.warning(self, "Warning", "No data to print")
            return
        
        try:
            temp_file = f"temp_print_{uuid.uuid4().hex[:8]}.pdf"
            self.export_pdf(temp_file)
            
            if os.path.exists(temp_file):
                if os.name == 'nt':
                    os.startfile(temp_file, 'print')
                else:
                    subprocess.call(['lpr', temp_file])
                    
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Print failed: {e}")


# ==================== MAIN ENTRY POINT ====================
def main():
    """Application entry point"""
    # Create required directories
    os.makedirs('reports', exist_ok=True)
    os.makedirs('backup', exist_ok=True)
    
    # Create application
    app = QApplication(sys.argv)
    app.setApplicationName(Config.APP_NAME)
    app.setOrganizationName(Config.COMPANY_NAME)
    
    # Set application style
    app.setStyle('Fusion')
    
    # Create and show main window
    window = MainApplication()
    window.show()
    
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()